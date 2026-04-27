from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from seed_normalizer import normalize_seed_workbook


class SeedNormalizerTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmpdir = tempfile.TemporaryDirectory()
        self._root = Path(self._tmpdir.name)
        self.source = self._root / "source.xlsx"
        self.target = self._root / "target.xlsx"
        self.report = self._root / "report.json"

    def tearDown(self) -> None:
        self._tmpdir.cleanup()

    def _write_source_workbook(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "inventory_items"
        ws.append(
            [
                "id",
                "asset_type",
                "asset_status",
                "key",
                "n_property_sn",
                "property_sn",
                "n_item_sn",
                "item_sn",
                "name",
                "name_code",
                "name_code2",
                "model",
                "specification",
                "unit",
                "count",
                "reserve_count",
                "purchase_date",
                "start_date",
                "due_date",
                "return_date",
                "location",
                "memo",
                "memo2",
                "keeper",
                "borrower",
                "created_at",
                "created_by",
                "updated_at",
                "updated_by",
                "deleted_at",
                "deleted_by",
                "condition_status",
            ]
        )
        ws.append(
            [
                1,
                "11",
                "0",
                "k1",
                "p1",
                "",
                "",
                "",
                "item-1",
                "2",
                "1",
                "M1",
                "",
                "台",
                "1",
                "7",
                "",
                "",
                "",
                "",
                "A倉",
                "",
                "",
                "keeper",
                "",
                "2026-01-01 00:00:00",
                "system",
                "",
                "",
                "",
                "",
                "",
            ]
        )
        ws.append(
            [
                2,
                "11",
                "0",
                "k2",
                "p2",
                "",
                "",
                "",
                "item-2",
                "40",
                "0",
                "M2",
                "",
                "台",
                "1",
                "0",
                "",
                "",
                "",
                "",
                "B倉",
                "",
                "",
                "keeper",
                "",
                "2026-01-01 00:00:00",
                "system",
                "",
                "",
                "",
                "",
                "",
            ]
        )

        cat = wb.create_sheet("asset_category_name")
        cat.append(["asset_category_name", "name_code", "name_code2", "description", ""])
        cat.append(["桌上型電腦", "02", "01", "一般用途", ""])

        status = wb.create_sheet("asset_status_code")
        status.append(["asset_status", "description", "", "", ""])
        status.append(["0", "庫存", "", "", ""])
        status.append(["1", "領用", "", "", ""])

        condition = wb.create_sheet("condition_status_code")
        condition.append(["condition_status", "description", "", "", ""])
        condition.append(["0", "良好", "", "", ""])

        wb.save(self.source)

    def test_normalize_seed_workbook_maps_and_backfills(self) -> None:
        self._write_source_workbook()
        report = normalize_seed_workbook(
            source_path=self.source,
            target_path=self.target,
            report_path=self.report,
        )

        self.assertEqual(report.inventory_rows_written, 2)
        self.assertIn("reserve_count", report.inventory_extra_columns)
        self.assertIn("create_at", report.inventory_missing_columns)
        self.assertEqual(report.backfilled_category_pairs, [["40", "00"]])

        wb = load_workbook(self.target)
        inv_ws = wb["inventory_items"]
        headers = [cell.value for cell in inv_ws[1]]
        first = list(next(inv_ws.iter_rows(min_row=2, max_row=2, values_only=True)))
        second = list(next(inv_ws.iter_rows(min_row=3, max_row=3, values_only=True)))
        row1 = {headers[idx]: first[idx] for idx in range(len(headers))}
        row2 = {headers[idx]: second[idx] for idx in range(len(headers))}

        self.assertEqual(row1["name_code"], "02")
        self.assertEqual(row1["name_code2"], "01")
        self.assertEqual(row2["name_code"], "40")
        self.assertEqual(row2["name_code2"], "00")
        self.assertIn(row1["create_at"], ("", None))
        self.assertIn(row1["update_at"], ("", None))

        cat_ws = wb["asset_category_name"]
        cat_headers = [cell.value for cell in cat_ws[1]]
        cat_rows = []
        for row in cat_ws.iter_rows(min_row=2, values_only=True):
            if not any(cell not in (None, "") for cell in row):
                continue
            cat_rows.append({cat_headers[idx]: row[idx] for idx in range(len(cat_headers))})
        pairs = {(str(row["name_code"]), str(row["name_code2"])) for row in cat_rows}
        self.assertIn(("02", "01"), pairs)
        self.assertIn(("40", "00"), pairs)

        backfilled = next(row for row in cat_rows if str(row["name_code"]) == "40" and str(row["name_code2"]) == "00")
        self.assertEqual(str(backfilled["description"]), "backfilled from seed normalization")

        status_ws = wb["asset_status_codes"]
        status_headers = [cell.value for cell in status_ws[1]]
        status_first = list(next(status_ws.iter_rows(min_row=2, max_row=2, values_only=True)))
        status_row = {status_headers[idx]: status_first[idx] for idx in range(len(status_headers))}
        self.assertEqual(str(status_row["code"]), "0")
        self.assertEqual(str(status_row["description"]), "庫存")

        self.assertTrue(self.report.exists())
        payload = json.loads(self.report.read_text(encoding="utf-8"))
        self.assertEqual(payload["inventory_rows_written"], 2)
        self.assertEqual(payload["backfilled_category_pairs"], [["40", "00"]])


if __name__ == "__main__":
    unittest.main()
