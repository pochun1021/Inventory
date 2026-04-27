import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

import db


class WorkbookResilienceTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmpdir = tempfile.TemporaryDirectory()
        self._original_db_path = db.DB_PATH
        self._original_lock_path = db.LOCK_PATH
        self._original_log_archive_dir = db.LOG_ARCHIVE_DIR

        db.DB_PATH = Path(self._tmpdir.name) / "inventory.xlsx"
        db.LOCK_PATH = Path(self._tmpdir.name) / "inventory.xlsx.lock"
        db.LOG_ARCHIVE_DIR = Path(self._tmpdir.name) / "log_archive"
        db.init_db()

    def tearDown(self) -> None:
        db.DB_PATH = self._original_db_path
        db.LOCK_PATH = self._original_lock_path
        db.LOG_ARCHIVE_DIR = self._original_log_archive_dir
        self._tmpdir.cleanup()

    def test_init_db_recovers_from_corrupted_workbook(self) -> None:
        db.DB_PATH.write_bytes(b"this-is-not-a-valid-xlsx")

        db.init_db()

        # Corrupted file should be preserved for troubleshooting.
        backups = list(db.DB_PATH.parent.glob("inventory.corrupt-*.xlsx"))
        self.assertEqual(len(backups), 1)
        self.assertTrue(backups[0].read_bytes().startswith(b"this-is-not-a-valid-xlsx"))

        # Rebuilt workbook should be readable and include expected sheets.
        wb = load_workbook(db.DB_PATH)
        self.assertIn("inventory_items", wb.sheetnames)
        self.assertIn("system_settings", wb.sheetnames)


if __name__ == "__main__":
    unittest.main()
