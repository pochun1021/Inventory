import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

import db


OLD_INVENTORY_HEADERS = [
    "id",
    "asset_type",
    "asset_status",
    "key",
    "n_property_sn",
    "property_sn",
    "n_item_sn",
    "item_sn",
    "name",
    "name_code",
    "name_code2",
    "model",
    "specification",
    "unit",
    "count",
    "purchase_date",
    "due_date",
    "return_date",
    "location",
    "memo",
    "memo2",
    "keeper",
    "created_at",
    "created_by",
    "updated_at",
    "updated_by",
    "deleted_at",
    "deleted_by",
]


class SchemaAlignmentTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmpdir = tempfile.TemporaryDirectory()
        self._original_db_path = db.DB_PATH
        self._original_lock_path = db.LOCK_PATH
        self._original_log_archive_dir = db.LOG_ARCHIVE_DIR
        self._original_asset_category_mapping_loaded = db.ASSET_CATEGORY_MAPPING_LOADED
        self._original_valid_name_code_pairs = set(db.VALID_NAME_CODE_PAIRS)

        db.DB_PATH = Path(self._tmpdir.name) / "inventory.xlsx"
        db.LOCK_PATH = Path(self._tmpdir.name) / "inventory.xlsx.lock"
        db.LOG_ARCHIVE_DIR = Path(self._tmpdir.name) / "log_archive"
        db.ASSET_CATEGORY_MAPPING_LOADED = True
        db.VALID_NAME_CODE_PAIRS.clear()

    def tearDown(self) -> None:
        db.DB_PATH = self._original_db_path
        db.LOCK_PATH = self._original_lock_path
        db.LOG_ARCHIVE_DIR = self._original_log_archive_dir
        db.ASSET_CATEGORY_MAPPING_LOADED = self._original_asset_category_mapping_loaded
        db.VALID_NAME_CODE_PAIRS.clear()
        db.VALID_NAME_CODE_PAIRS.update(self._original_valid_name_code_pairs)
        self._tmpdir.cleanup()

    def _write_old_inventory_workbook(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "inventory_items"
        ws.append(OLD_INVENTORY_HEADERS)
        ws.append(
            [
                1,
                "A1",
                "0",
                "tmp-20260422-0001",
                "",
                "",
                "ABC123",
                "",
                "測試品項",
                "",
                "",
                "M1",
                "規格",
                "個",
                "1",
                "2026/04/22",
                "",
                "",
                "A倉",
                "",
                "",
                "管理員",
                "2026-04-22 10:00:00",
                "system",
                "",
                "",
                "",
                "",
            ]
        )
        wb.save(db.DB_PATH)

    def test_init_db_migrates_inventory_headers_without_data_loss(self) -> None:
        self._write_old_inventory_workbook()

        db.init_db()

        wb = load_workbook(db.DB_PATH)
        ws = wb["inventory_items"]
        headers = [cell.value for cell in ws[1]]
        row = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
        row_map = {headers[idx]: row[idx] for idx in range(len(headers))}

        self.assertEqual(headers, db.SHEETS["inventory_items"])
        self.assertEqual(row_map["id"], 1)
        self.assertEqual(row_map["name"], "測試品項")
        self.assertEqual(row_map["asset_status"], "0")
        self.assertEqual(row_map["keeper"], "管理員")
        self.assertIn(row_map["condition_status"], ("", None))
        self.assertIn(row_map["borrower"], ("", None))
        self.assertIn(row_map["start_date"], ("", None))
        self.assertIn(row_map["create_at"], ("", None))
        self.assertIn(row_map["update_at"], ("", None))

    def test_init_db_creates_pos_alignment_sheets_and_seeds(self) -> None:
        self._write_old_inventory_workbook()

        db.init_db()

        wb = load_workbook(db.DB_PATH)
        self.assertIn("condition_status_code", wb.sheetnames)
        self.assertIn("asset_category_name", wb.sheetnames)
        self.assertIn("inventory_items_schema", wb.sheetnames)
        self.assertIn("transaction_log", wb.sheetnames)

        condition_rows = db._read_rows(wb["condition_status_code"])  # noqa: SLF001
        self.assertEqual(
            {(str(row["condition_status"]), str(row["description"])) for row in condition_rows},
            {("0", "良好"), ("1", "損壞"), ("2", "報廢")},
        )

        schema_rows = db._read_rows(wb["inventory_items_schema"])  # noqa: SLF001
        schema_fields = {str(row["field"]) for row in schema_rows}
        self.assertIn("condition_status", schema_fields)
        self.assertIn("borrower", schema_fields)
        self.assertIn("create_at", schema_fields)

        category_rows = db._read_rows(wb["asset_category_name"])  # noqa: SLF001
        category_pairs = {
            (str(row["name_code"]), str(row["name_code2"]))
            for row in category_rows
            if row.get("name_code") and row.get("name_code2")
        }
        self.assertIn(("01", "01"), category_pairs)

        transaction_headers = [cell.value for cell in wb["transaction_log"][1]]
        self.assertEqual(transaction_headers, db.SHEETS["transaction_log"])


if __name__ == "__main__":
    unittest.main()
