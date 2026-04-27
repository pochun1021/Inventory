from __future__ import annotations

import argparse
import hashlib
import json
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import db
from migration_service import TARGET_TABLES, run_xlsx_to_supabase_migration
from supabase_client import get_supabase_client

REPORT_DIR = Path(__file__).resolve().parent / "migration_reports"
PAGE_SIZE = 1000
PRIMARY_KEY_COLUMNS: dict[str, tuple[str, ...]] = {
    "asset_status_codes": ("code",),
    "condition_status_code": ("condition_status",),
    "asset_category_name": ("name_code", "name_code2"),
    "inventory_items": ("id",),
    "issue_requests": ("id",),
    "issue_items": ("id",),
    "borrow_requests": ("id",),
    "borrow_request_lines": ("id",),
    "borrow_allocations": ("id",),
    "donation_requests": ("id",),
    "donation_items": ("id",),
    "movement_ledger": ("id",),
    "operation_logs": ("id",),
    "order_sn": ("name",),
}


@dataclass
class TableDiff:
    table: str
    xlsx_rows: int
    supabase_rows: int
    xlsx_digest: str
    supabase_digest: str
    status: str


def _xlsx_rows(table: str) -> list[dict[str, Any]]:
    if not db.DB_PATH.exists():
        return []
    wb = load_workbook(db.DB_PATH)
    if table not in wb.sheetnames:
        return []
    raw_rows = db._read_rows(wb[table])  # noqa: SLF001
    if table == "asset_category_name" and "inventory_items" in wb.sheetnames:
        inventory_rows = db._read_rows(wb["inventory_items"])  # noqa: SLF001
        for row in inventory_rows:
            name_code = _normalize_value(row.get("name_code"))
            name_code2 = _normalize_value(row.get("name_code2"))
            if name_code is None or name_code2 is None:
                continue
            raw_rows.append(
                {
                    "name_code": name_code,
                    "asset_category_name": None,
                    "name_code2": name_code2,
                    "description": "backfilled from inventory_items during migration",
                    "created_at": "",
                    "updated_at": "",
                }
            )
    return _canonicalize_rows(table, raw_rows)


def _supabase_rows(table: str) -> list[dict[str, Any]]:
    client = get_supabase_client()
    offset = 0
    rows: list[dict[str, Any]] = []
    while True:
        response = client.table(table).select("*").range(offset, offset + PAGE_SIZE - 1).execute()
        batch = response.data or []
        if not batch:
            break
        rows.extend(batch)
        if len(batch) < PAGE_SIZE:
            break
        offset += PAGE_SIZE
    return _canonicalize_rows(table, rows)


def _normalize_value(value: Any) -> Any:
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned if cleaned != "" else None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bool):
        return str(value).lower()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, dict) or isinstance(value, list):
        return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    return value


def _canonicalize_rows(table: str, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    columns = db.SHEETS.get(table)
    if columns:
        normalized_rows = [{column: _normalize_value(row.get(column)) for column in columns} for row in rows]
    else:
        normalized_rows = [{key: _normalize_value(value) for key, value in row.items()} for row in rows]

    primary_keys = PRIMARY_KEY_COLUMNS.get(table, ())
    if not primary_keys:
        return normalized_rows

    deduped_by_pk: dict[tuple[Any, ...], dict[str, Any]] = {}
    passthrough_rows: list[dict[str, Any]] = []
    for row in normalized_rows:
        key = tuple(row.get(column) for column in primary_keys)
        if any(value is None for value in key):
            passthrough_rows.append(row)
            continue
        deduped_by_pk[key] = row

    return passthrough_rows + list(deduped_by_pk.values())


def _digest_rows(rows: list[dict[str, Any]]) -> str:
    normalized = [json.dumps(row, ensure_ascii=False, sort_keys=True, default=str) for row in rows]
    normalized.sort()
    payload = "\n".join(normalized).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _collect_table_results() -> list[TableDiff]:
    table_results: list[TableDiff] = []
    for table in TARGET_TABLES:
        xlsx_rows = _xlsx_rows(table)
        supabase_rows = _supabase_rows(table)
        xlsx_digest = _digest_rows(xlsx_rows)
        supabase_digest = _digest_rows(supabase_rows)
        status = "match" if xlsx_digest == supabase_digest else "mismatch"
        table_results.append(
            TableDiff(
                table=table,
                xlsx_rows=len(xlsx_rows),
                supabase_rows=len(supabase_rows),
                xlsx_digest=xlsx_digest,
                supabase_digest=supabase_digest,
                status=status,
            )
        )
    return table_results


def _ensure_xlsx_source() -> None:
    if not db.DB_PATH.exists():
        raise FileNotFoundError(f"XLSX source not found: {db.DB_PATH}")

    wb = load_workbook(db.DB_PATH)
    available_sheets = set(wb.sheetnames)
    expected_sheets = set(TARGET_TABLES)
    if not (available_sheets & expected_sheets):
        raise RuntimeError("XLSX source does not contain any reconciliation target sheets")


def reconcile(*, repair: bool, repair_tables: list[str] | None = None) -> dict[str, Any]:
    _ensure_xlsx_source()
    table_results = _collect_table_results()
    mismatches = [item for item in table_results if item.status == "mismatch"]
    repaired = False
    repair_report: dict[str, Any] | None = None
    effective_repair = repair or bool(repair_tables)

    if effective_repair and mismatches:
        repair_report = run_xlsx_to_supabase_migration(
            dry_run=False,
            replace_existing=True,
            target_tables=repair_tables,
        )
        repaired = repair_report.get("status") == "success"
        table_results = _collect_table_results()
        mismatches = [item for item in table_results if item.status == "mismatch"]

    status = "success" if not mismatches else "mismatch"
    if effective_repair and repaired and not mismatches:
        status = "repaired"
    report = {
        "status": status,
        "table_results": [asdict(item) for item in table_results],
        "mismatch_tables": [item.table for item in mismatches],
        "repaired": repaired,
        "repair_report": repair_report,
    }

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    report_path = REPORT_DIR / "reconcile-latest.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description="Compare XLSX and Supabase datasets and optionally repair by re-syncing.")
    parser.add_argument("--repair", action="store_true", help="Run XLSX -> Supabase full migration when mismatch is found")
    parser.add_argument(
        "--repair-table",
        action="append",
        dest="repair_tables",
        help="Limit repair scope to specific table(s); can be used multiple times",
    )
    args = parser.parse_args()
    repair_tables = args.repair_tables or None
    if repair_tables:
        unknown_tables = sorted(set(repair_tables) - set(TARGET_TABLES))
        if unknown_tables:
            parser.error(f"Unknown --repair-table value(s): {', '.join(unknown_tables)}")

    report = reconcile(repair=args.repair, repair_tables=repair_tables)
    print(json.dumps(report, ensure_ascii=False, indent=2))
    if report["status"] == "mismatch":
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
