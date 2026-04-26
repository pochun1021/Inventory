from __future__ import annotations

import json
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import db
from supabase_client import get_supabase_client

REPORT_DIR = Path(__file__).resolve().parent / "migration_reports"
BATCH_SIZE = 500

ID_COLUMNS: dict[str, set[str]] = {
    "inventory_items": {"id", "donation_request_id", "parent_item_id"},
    "asset_status_codes": set(),
    "condition_status_code": set(),
    "asset_category_name": set(),
    "issue_requests": {"id"},
    "issue_items": {"id", "request_id", "item_id", "quantity"},
    "borrow_requests": {"id"},
    "borrow_request_lines": {"id", "request_id", "requested_qty"},
    "borrow_allocations": {"id", "request_id", "line_id", "item_id"},
    "donation_requests": {"id"},
    "donation_items": {"id", "request_id", "item_id", "quantity"},
    "movement_ledger": {"id", "item_id", "entity_id"},
    "operation_logs": {"id", "entity_id"},
    "order_sn": {"current_value"},
    "sync_job_log": {"id", "total_rows"},
}

TARGET_TABLES: list[str] = [
    "asset_status_codes",
    "condition_status_code",
    "asset_category_name",
    "inventory_items",
    "issue_requests",
    "issue_items",
    "borrow_requests",
    "borrow_request_lines",
    "borrow_allocations",
    "donation_requests",
    "donation_items",
    "movement_ledger",
    "operation_logs",
    "order_sn",
]

ORPHAN_ITEM_SKIP_RULES: dict[str, str] = {
    "issue_items": "item_id",
    "donation_items": "item_id",
}
MAX_SKIPPED_SAMPLE_SIZE = 20


@dataclass
class MigrationTableResult:
    table: str
    source_rows: int
    migrated_rows: int
    skipped_rows: int
    status: str
    message: str = ""
    skip_reason: str = ""
    skipped_samples: list[int] = field(default_factory=list)


@dataclass
class MigrationReport:
    job_id: str
    started_at: str
    finished_at: str = ""
    dry_run: bool = False
    status: str = "running"
    tables: list[MigrationTableResult] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def _now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _normalize_row(table: str, row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    id_columns = ID_COLUMNS.get(table, set())
    for key, value in row.items():
        if value is None:
            normalized[key] = None
            continue
        if isinstance(value, str):
            cleaned = value.strip()
            if cleaned == "":
                normalized[key] = None
                continue
            if key in id_columns:
                try:
                    normalized[key] = int(cleaned)
                except ValueError:
                    normalized[key] = None
            else:
                normalized[key] = cleaned
            continue
        normalized[key] = value
    return normalized


def _chunks(items: list[dict[str, Any]], size: int) -> list[list[dict[str, Any]]]:
    return [items[i : i + size] for i in range(0, len(items), size)]


def _read_sheet_rows(table: str) -> list[dict[str, Any]]:
    if not db.DB_PATH.exists():
        return []
    wb = load_workbook(db.DB_PATH)
    if table not in wb.sheetnames:
        return []
    return db._read_rows(wb[table])  # noqa: SLF001


def _write_report(report: MigrationReport) -> Path:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    path = REPORT_DIR / f"{report.job_id}.json"
    path.write_text(json.dumps(asdict(report), ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def get_migration_report(job_id: str) -> dict[str, Any] | None:
    report_path = REPORT_DIR / f"{job_id}.json"
    if not report_path.exists():
        return None
    return json.loads(report_path.read_text(encoding="utf-8"))


def _adjust_sequences() -> None:
    client = get_supabase_client()
    # Requires schema function from backend/supabase_sql/schema.sql.
    client.rpc("admin_set_sequences", {}).execute()


def _extract_valid_item_ids(rows: list[dict[str, Any]]) -> set[int]:
    valid_ids: set[int] = set()
    for row in rows:
        value = row.get("id")
        if isinstance(value, int) and value > 0:
            valid_ids.add(value)
    return valid_ids


def _filter_orphan_item_rows(
    table: str,
    rows: list[dict[str, Any]],
    valid_item_ids: set[int],
) -> tuple[list[dict[str, Any]], int, list[int], str]:
    item_id_field = ORPHAN_ITEM_SKIP_RULES.get(table)
    if not item_id_field:
        return rows, 0, [], ""

    kept_rows: list[dict[str, Any]] = []
    skipped_count = 0
    skipped_samples: list[int] = []
    skipped_sample_set: set[int] = set()

    for row in rows:
        item_id = row.get(item_id_field)
        if isinstance(item_id, int) and item_id > 0 and item_id not in valid_item_ids:
            skipped_count += 1
            if len(skipped_samples) < MAX_SKIPPED_SAMPLE_SIZE and item_id not in skipped_sample_set:
                skipped_sample_set.add(item_id)
                skipped_samples.append(item_id)
            continue
        kept_rows.append(row)

    skip_reason = "orphan_item_id" if skipped_count > 0 else ""
    return kept_rows, skipped_count, skipped_samples, skip_reason


def run_xlsx_to_supabase_migration(*, dry_run: bool) -> dict[str, Any]:
    job_id = datetime.now().strftime("%Y%m%d%H%M%S%f")
    report = MigrationReport(job_id=job_id, started_at=_now_str(), dry_run=dry_run)
    client = get_supabase_client()

    try:
        valid_item_ids: set[int] = set()
        for table in TARGET_TABLES:
            rows = _read_sheet_rows(table)
            normalized_rows = [_normalize_row(table, row) for row in rows]
            payload_rows = [row for row in normalized_rows if any(value is not None for value in row.values())]
            empty_skipped_rows = len(normalized_rows) - len(payload_rows)

            payload_rows, orphan_skipped_rows, skipped_samples, skip_reason = _filter_orphan_item_rows(
                table,
                payload_rows,
                valid_item_ids,
            )
            skipped_rows = empty_skipped_rows + orphan_skipped_rows
            table_status = "ok_with_skips" if skipped_rows > 0 else "ok"
            table_message = ""
            if orphan_skipped_rows > 0:
                table_message = f"skipped {orphan_skipped_rows} rows with orphan item references"

            if table == "inventory_items":
                valid_item_ids = _extract_valid_item_ids(payload_rows)

            if dry_run:
                report.tables.append(
                    MigrationTableResult(
                        table=table,
                        source_rows=len(rows),
                        migrated_rows=0,
                        skipped_rows=skipped_rows,
                        status="dry_run",
                        message=table_message,
                        skip_reason=skip_reason,
                        skipped_samples=skipped_samples,
                    )
                )
                continue

            migrated = 0
            if payload_rows:
                for batch in _chunks(payload_rows, BATCH_SIZE):
                    response = client.table(table).upsert(batch).execute()
                    data = response.data or []
                    migrated += len(data) if isinstance(data, list) else len(batch)

            report.tables.append(
                MigrationTableResult(
                    table=table,
                    source_rows=len(rows),
                    migrated_rows=migrated,
                    skipped_rows=skipped_rows,
                    status=table_status,
                    message=table_message,
                    skip_reason=skip_reason,
                    skipped_samples=skipped_samples,
                )
            )

        if not dry_run:
            _adjust_sequences()

        report.status = "success"
    except Exception as exc:  # pragma: no cover - external service behavior
        report.status = "failed"
        report.errors.append(str(exc))
    finally:
        report.finished_at = _now_str()
        _write_report(report)

    return asdict(report)
