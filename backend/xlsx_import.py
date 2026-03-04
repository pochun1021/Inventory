from datetime import date, datetime
from io import BytesIO
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook
from pydantic import ValidationError


REQUIRED_HEADERS = [
    "備註",
    "規格(大小/容量)",
    "財產編號",
    "品名",
    "型號",
    "單位",
    "購置日期",
    "放置地點",
    "保管人（單位）",
]


def import_inventory_items_from_xlsx_content(
    file_content: bytes,
    item_create_model,
    to_db_payload,
    create_item,
    selected_kind: str,
) -> dict[str, Any]:
    if not file_content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    try:
        workbook = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Invalid xlsx file: {exc}") from exc

    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if header_row is None:
        raise HTTPException(status_code=400, detail="XLSX does not contain a header row")

    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    missing_headers = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing_headers:
        raise HTTPException(status_code=400, detail=f"Missing required headers: {', '.join(missing_headers)}")

    total = 0
    created = 0
    errors: list[dict[str, Any]] = []

    for row_index, row in enumerate(rows, start=2):
        if row is None:
            continue

        row_data = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
        if all(value in (None, "") for value in row_data.values()):
            continue

        total += 1
        try:
            payload = {
                "類別": selected_kind,
                "備註": str(row_data.get("備註") or "").strip(),
                "規格(大小/容量)": str(row_data.get("規格(大小/容量)") or "").strip(),
                "財產編號": str(row_data.get("財產編號") or "").strip(),
                "品名": str(row_data.get("品名") or "").strip(),
                "型號": str(row_data.get("型號") or "").strip(),
                "單位": str(row_data.get("單位") or "").strip(),
                "購置日期": normalize_purchase_date(row_data.get("購置日期")),
                "放置地點": str(row_data.get("放置地點") or "").strip(),
                "保管人（單位）": str(row_data.get("保管人（單位）") or "").strip(),
            }
            item = item_create_model.model_validate(payload)
            create_item(to_db_payload(item))
            created += 1
        except ValidationError as exc:
            errors.append({"row": row_index, "message": exc.errors()[0]["msg"]})
        except Exception as exc:  # noqa: BLE001
            errors.append({"row": row_index, "message": str(exc)})

    return {
        "total": total,
        "created": created,
        "failed": total - created,
        "errors": errors,
    }


def normalize_purchase_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()

