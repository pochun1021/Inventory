from __future__ import annotations

import argparse
import json
import shutil
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

import db


def _to_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_code(value: Any) -> str:
    raw = _to_str(value)
    if not raw:
        return ""
    if raw.isdigit() and len(raw) <= 2:
        return raw.zfill(2)
    return raw


def _read_rows(ws) -> list[dict[str, Any]]:
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]] if ws.max_row >= 1 else []
    rows: list[dict[str, Any]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row[header] = raw[idx] if idx < len(raw) else None
        if any(_to_str(value) != "" for value in row.values()):
            rows.append(row)
    return rows


def _create_target_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, headers in db.SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
    return wb


@dataclass
class NormalizationReport:
    source_path: str
    target_path: str
    backup_path: str
    generated_at: str
    inventory_rows_written: int
    category_rows_written: int
    status_rows_written: int
    condition_rows_written: int
    inventory_missing_columns: list[str]
    inventory_extra_columns: list[str]
    backfilled_category_pairs: list[list[str]]


def normalize_seed_workbook(
    *,
    source_path: Path,
    target_path: Path,
    report_path: Path,
) -> NormalizationReport:
    source_wb = load_workbook(source_path)
    target_wb = _create_target_workbook()

    source_inventory_rows: list[dict[str, Any]] = []
    source_inventory_headers: list[str] = []
    if "inventory_items" in source_wb.sheetnames:
        ws = source_wb["inventory_items"]
        source_inventory_headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in ws[1]
        ] if ws.max_row >= 1 else []
        source_inventory_rows = _read_rows(ws)

    target_inventory_ws = target_wb["inventory_items"]
    target_inventory_headers = db.SHEETS["inventory_items"]

    target_inventory_rows: list[dict[str, Any]] = []
    observed_pairs: set[tuple[str, str]] = set()
    for row in source_inventory_rows:
        mapped: dict[str, Any] = {}
        for header in target_inventory_headers:
            mapped[header] = row.get(header, "")

        mapped["name_code"] = _normalize_code(mapped.get("name_code"))
        mapped["name_code2"] = _normalize_code(mapped.get("name_code2"))
        if (mapped["name_code"] and not mapped["name_code2"]) or (not mapped["name_code"] and mapped["name_code2"]):
            mapped["name_code"] = ""
            mapped["name_code2"] = ""

        pair = (mapped["name_code"], mapped["name_code2"])
        if pair[0] and pair[1]:
            observed_pairs.add(pair)
        target_inventory_rows.append(mapped)

    for row in target_inventory_rows:
        target_inventory_ws.append([row.get(header, "") for header in target_inventory_headers])

    source_category_rows: list[dict[str, Any]] = []
    if "asset_category_name" in source_wb.sheetnames:
        source_category_rows = _read_rows(source_wb["asset_category_name"])

    target_category_ws = target_wb["asset_category_name"]
    category_pairs: set[tuple[str, str]] = set()
    category_rows_out: list[dict[str, Any]] = []
    for row in source_category_rows:
        name_code = _normalize_code(row.get("name_code"))
        name_code2 = _normalize_code(row.get("name_code2"))
        if not name_code or not name_code2:
            continue
        pair = (name_code, name_code2)
        if pair in category_pairs:
            continue
        category_pairs.add(pair)
        category_rows_out.append(
            {
                "name_code": name_code,
                "asset_category_name": _to_str(row.get("asset_category_name")),
                "name_code2": name_code2,
                "description": _to_str(row.get("description")),
            }
        )

    backfilled_pairs: list[list[str]] = []
    for pair in sorted(observed_pairs):
        if pair in category_pairs:
            continue
        category_pairs.add(pair)
        backfilled_pairs.append([pair[0], pair[1]])
        category_rows_out.append(
            {
                "name_code": pair[0],
                "asset_category_name": "",
                "name_code2": pair[1],
                "description": "backfilled from seed normalization",
            }
        )

    for row in sorted(category_rows_out, key=lambda item: (item["name_code"], item["name_code2"])):
        target_category_ws.append(
            [
                row["name_code"],
                row["asset_category_name"],
                row["name_code2"],
                row["description"],
            ]
        )

    source_status_rows: list[dict[str, Any]] = []
    if "asset_status_code" in source_wb.sheetnames:
        source_status_rows = _read_rows(source_wb["asset_status_code"])
    elif "asset_status_codes" in source_wb.sheetnames:
        source_status_rows = _read_rows(source_wb["asset_status_codes"])

    target_status_ws = target_wb["asset_status_codes"]
    status_rows_out: list[dict[str, str]] = []
    seen_status_codes: set[str] = set()
    for row in source_status_rows:
        code = _to_str(row.get("code") or row.get("asset_status"))
        if not code or code in seen_status_codes:
            continue
        seen_status_codes.add(code)
        status_rows_out.append(
            {
                "code": code,
                "description": _to_str(row.get("description")),
                "created_at": _to_str(row.get("created_at")),
                "updated_at": _to_str(row.get("updated_at")),
            }
        )
    for row in status_rows_out:
        target_status_ws.append([row["code"], row["description"], row["created_at"], row["updated_at"]])

    source_condition_rows: list[dict[str, Any]] = []
    if "condition_status_code" in source_wb.sheetnames:
        source_condition_rows = _read_rows(source_wb["condition_status_code"])

    target_condition_ws = target_wb["condition_status_code"]
    condition_rows_out: list[dict[str, str]] = []
    seen_conditions: set[str] = set()
    for row in source_condition_rows:
        code = _to_str(row.get("condition_status"))
        if not code or code in seen_conditions:
            continue
        seen_conditions.add(code)
        condition_rows_out.append(
            {
                "condition_status": code,
                "description": _to_str(row.get("description")),
            }
        )
    for row in condition_rows_out:
        target_condition_ws.append([row["condition_status"], row["description"]])

    had_existing_target = target_path.exists()
    backup_path = target_path.with_name(f"{target_path.stem}.backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}{target_path.suffix}")
    if had_existing_target:
        shutil.copy2(target_path, backup_path)

    target_wb.save(target_path)

    missing_columns = [col for col in target_inventory_headers if col not in source_inventory_headers]
    extra_columns = [col for col in source_inventory_headers if col and col not in target_inventory_headers]

    report = NormalizationReport(
        source_path=str(source_path),
        target_path=str(target_path),
        backup_path=str(backup_path) if had_existing_target else "",
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        inventory_rows_written=len(target_inventory_rows),
        category_rows_written=len(category_rows_out),
        status_rows_written=len(status_rows_out),
        condition_rows_written=len(condition_rows_out),
        inventory_missing_columns=missing_columns,
        inventory_extra_columns=extra_columns,
        backfilled_category_pairs=backfilled_pairs,
    )

    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(asdict(report), ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Normalize source XLSX into backend seed workbook format.")
    parser.add_argument(
        "--source",
        default=str(Path.home() / "inventory_db.xlsx"),
        help="Source XLSX path",
    )
    parser.add_argument(
        "--target",
        default=str(db.DB_PATH),
        help="Target normalized XLSX path",
    )
    parser.add_argument(
        "--report",
        default=str(Path(__file__).resolve().parent / "migration_reports" / "seed-normalization-latest.json"),
        help="Normalization report output path",
    )
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    report = normalize_seed_workbook(
        source_path=Path(args.source).expanduser(),
        target_path=Path(args.target).expanduser(),
        report_path=Path(args.report).expanduser(),
    )
    print(json.dumps(asdict(report), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
