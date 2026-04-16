from __future__ import annotations

import json
import re
from contextlib import contextmanager
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from filelock import FileLock
from openpyxl import Workbook, load_workbook

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "inventory.xlsx"
LOCK_PATH = BASE_DIR / "inventory.xlsx.lock"
ASSET_CATEGORY_NAME_PATH = BASE_DIR.parent / "asset_category_name.xlsx"
LOG_ARCHIVE_DIR = BASE_DIR / "log_archive"
HOT_LOG_RETENTION_DAYS = 90
MAX_BORROW_RESERVATION_DAYS = 30
BORROW_RESERVATION_CANCEL_AFTER_DAYS = 3
ARCHIVE_FILE_PATTERN = re.compile(r"^logs_(\d{6})\.xlsx$")

SHEETS: dict[str, list[str]] = {
    "inventory_items": [
        "id",
        "asset_type",
        "asset_status",
        "key",
        "n_property_sn",
        "property_sn",
        "n_item_sn",
        "item_sn",
        "name",
        "name_code",
        "name_code2",
        "model",
        "specification",
        "unit",
        "count",
        "purchase_date",
        "due_date",
        "return_date",
        "location",
        "memo",
        "memo2",
        "keeper",
        "created_at",
        "created_by",
        "updated_at",
        "updated_by",
        "deleted_at",
        "deleted_by",
    ],
    "order_sn": [
        "name",
        "current_value",
    ],
    "asset_status_codes": [
        "code",
        "description",
        "created_at",
        "updated_at",
    ],
    "operation_logs": [
        "id",
        "action",
        "entity",
        "entity_id",
        "status",
        "detail",
        "created_at",
    ],
    "movement_ledger": [
        "id",
        "item_id",
        "from_status",
        "to_status",
        "action",
        "entity",
        "entity_id",
        "operator",
        "created_at",
    ],
    "issue_requests": [
        "id",
        "requester",
        "department",
        "purpose",
        "request_date",
        "memo",
        "created_at",
    ],
    "issue_items": [
        "id",
        "request_id",
        "item_id",
        "quantity",
        "note",
    ],
    "borrow_requests": [
        "id",
        "borrower",
        "department",
        "purpose",
        "borrow_date",
        "due_date",
        "return_date",
        "status",
        "memo",
        "created_at",
    ],
    "borrow_items": [
        "id",
        "request_id",
        "item_id",
        "quantity",
        "note",
    ],
    "borrow_request_lines": [
        "id",
        "request_id",
        "item_name",
        "item_model",
        "requested_qty",
        "note",
    ],
    "borrow_allocations": [
        "id",
        "request_id",
        "line_id",
        "item_id",
        "note",
    ],
    "donation_requests": [
        "id",
        "donor",
        "department",
        "recipient",
        "purpose",
        "donation_date",
        "memo",
        "created_at",
    ],
    "donation_items": [
        "id",
        "request_id",
        "item_id",
        "quantity",
        "note",
    ],
}

STRING_FIELDS: dict[str, list[str]] = {
    "inventory_items": [
        "asset_type",
        "asset_status",
        "key",
        "n_property_sn",
        "property_sn",
        "n_item_sn",
        "item_sn",
        "name",
        "name_code",
        "name_code2",
        "model",
        "specification",
        "unit",
        "count",
        "purchase_date",
        "due_date",
        "return_date",
        "location",
        "memo",
        "memo2",
        "keeper",
        "created_at",
        "created_by",
        "updated_at",
        "updated_by",
        "deleted_at",
        "deleted_by",
    ],
    "issue_requests": ["requester", "department", "purpose", "request_date", "memo", "created_at"],
    "issue_items": ["note"],
    "borrow_requests": [
        "borrower",
        "department",
        "purpose",
        "borrow_date",
        "due_date",
        "return_date",
        "status",
        "memo",
        "created_at",
    ],
    "borrow_items": ["note"],
    "borrow_request_lines": ["item_name", "item_model", "note"],
    "borrow_allocations": ["note"],
    "donation_requests": [
        "donor",
        "department",
        "recipient",
        "purpose",
        "donation_date",
        "memo",
        "created_at",
    ],
    "donation_items": ["note"],
    "asset_status_codes": ["code", "description", "created_at", "updated_at"],
    "movement_ledger": ["from_status", "to_status", "action", "entity", "operator", "created_at"],
}

REMOVED_SHEETS = {
    "stock_balances",
    "stock_movements",
}


def _now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _date_sn() -> str:
    return datetime.now().strftime("%Y%m%d")


def _to_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _to_str(value: Any, default: str = "") -> str:
    if value is None:
        return default
    return value if isinstance(value, str) else str(value)


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _has_cjk(value: Any) -> bool:
    if value is None:
        return False
    return any("\u4e00" <= char <= "\u9fff" for char in str(value))


KIND_TO_ASSET_TYPE = {
    "asset": "11",
    "item": "A1",
    "other": "A2",
}
DEFAULT_ASSET_STATUS_CODES: list[tuple[str, str]] = [
    ("0", "庫存"),
    ("1", "領用"),
    ("2", "借用"),
    ("3", "捐贈"),
    ("4", "減損"),
    ("5", "報廢"),
]
ASSET_TYPE_TO_KIND = {value: key for key, value in KIND_TO_ASSET_TYPE.items()}
VALID_NAME_CODE_PAIRS: set[tuple[str, str]] = set()
ASSET_CATEGORY_MAPPING_LOADED = False


def _normalize_asset_type_input(asset_type: Any) -> str:
    raw = _to_str(asset_type).strip()
    if raw in ASSET_TYPE_TO_KIND:
        return raw
    return KIND_TO_ASSET_TYPE.get(raw, "A2")


def _asset_type_to_order_sn_name(asset_type: Any) -> str:
    normalized = _normalize_asset_type_input(asset_type)
    return ASSET_TYPE_TO_KIND.get(normalized, "other")


def _normalize_name_code_value(value: Any) -> str:
    raw = _to_str(value).strip()
    if not raw:
        return ""
    if raw.isdigit() and len(raw) <= 2:
        return raw.zfill(2)
    return raw


def _normalize_name_codes(name_code: Any, name_code2: Any) -> tuple[str, str]:
    normalized_code = _normalize_name_code_value(name_code)
    normalized_code2 = _normalize_name_code_value(name_code2)
    if not normalized_code or not normalized_code2:
        return "", ""
    if not VALID_NAME_CODE_PAIRS:
        return "", ""
    if (normalized_code, normalized_code2) not in VALID_NAME_CODE_PAIRS:
        return "", ""
    return normalized_code, normalized_code2


def _load_asset_category_mapping() -> None:
    global ASSET_CATEGORY_MAPPING_LOADED
    if ASSET_CATEGORY_MAPPING_LOADED:
        return

    pairs: set[tuple[str, str]] = set()
    if ASSET_CATEGORY_NAME_PATH.exists():
        try:
            wb = load_workbook(ASSET_CATEGORY_NAME_PATH, read_only=True, data_only=True)
            ws = wb["asset_category_name"] if "asset_category_name" in wb.sheetnames else wb.active
            rows = ws.iter_rows(min_row=2, values_only=True)
            for row in rows:
                if not row:
                    continue
                name_code = _normalize_name_code_value(row[0] if len(row) > 0 else "")
                name_code2 = _normalize_name_code_value(row[1] if len(row) > 1 else "")
                if not name_code or not name_code2:
                    continue
                pairs.add((name_code, name_code2))
        except Exception:  # noqa: BLE001
            pairs = set()

    VALID_NAME_CODE_PAIRS.clear()
    VALID_NAME_CODE_PAIRS.update(pairs)
    ASSET_CATEGORY_MAPPING_LOADED = True


def _inventory_property_number(row: dict[str, Any]) -> str:
    for field in ("n_property_sn", "property_sn", "n_item_sn", "item_sn"):
        value = _to_str(row.get(field)).strip()
        if value:
            return value
    return ""


def _inventory_key(row: dict[str, Any], property_number: str) -> str:
    key = _to_str(row.get("key")).strip()
    if key:
        return key
    if property_number:
        return property_number
    item_id = _to_int(row.get("id"))
    return f"item-{item_id}" if item_id > 0 else ""


def _to_inventory_create_row(new_id: int, item_data: dict[str, Any], property_number: str) -> dict[str, Any]:
    asset_type = _normalize_asset_type_input(item_data.get("asset_type"))
    n_property_sn = _to_str(item_data.get("n_property_sn")).strip()
    property_sn = _to_str(item_data.get("property_sn")).strip()
    n_item_sn = _to_str(item_data.get("n_item_sn")).strip()
    item_sn = _to_str(item_data.get("item_sn")).strip()
    if not any((n_property_sn, property_sn, n_item_sn, item_sn)) and property_number:
        if asset_type == "11":
            n_property_sn = property_number
        else:
            n_item_sn = property_number
    row_for_key = {
        "id": new_id,
        "key": _to_str(item_data.get("key")).strip(),
        "n_property_sn": n_property_sn,
        "property_sn": property_sn,
        "n_item_sn": n_item_sn,
        "item_sn": item_sn,
    }
    serial_for_key = _inventory_property_number(row_for_key)
    name_code, name_code2 = _normalize_name_codes(item_data.get("name_code"), item_data.get("name_code2"))
    now = _now_str()
    count = 1
    row = {
        "id": new_id,
        "asset_type": asset_type,
        "asset_status": _to_str(item_data.get("asset_status")).strip() or "0",
        "key": _inventory_key(row_for_key, serial_for_key),
        "n_property_sn": n_property_sn,
        "property_sn": property_sn,
        "n_item_sn": n_item_sn,
        "item_sn": item_sn,
        "name": _to_str(item_data.get("name")),
        "name_code": name_code,
        "name_code2": name_code2,
        "model": _to_str(item_data.get("model")),
        "specification": _to_str(item_data.get("specification")),
        "unit": _to_str(item_data.get("unit")),
        "count": str(count),
        "purchase_date": _to_str(item_data.get("purchase_date")),
        "due_date": _to_str(item_data.get("due_date")),
        "return_date": _to_str(item_data.get("return_date")),
        "location": _to_str(item_data.get("location")),
        "memo": _to_str(item_data.get("memo")),
        "memo2": _to_str(item_data.get("memo2")),
        "keeper": _to_str(item_data.get("keeper")),
        "created_at": now,
        "created_by": "system",
        "updated_at": "",
        "updated_by": "",
        "deleted_at": "",
        "deleted_by": "",
    }
    return row


def _to_inventory_api_row(
    row: dict[str, Any],
    *,
    donation_map: dict[int, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    item_id = _to_int(row.get("id"))
    donation_info = donation_map.get(item_id, {}) if donation_map else {}
    donated_at = _to_str(donation_info.get("donated_at"))
    donation_request_id = _to_int(donation_info.get("donation_request_id"))
    if not donated_at and _to_str(row.get("asset_status")) == "3":
        donated_at = _to_str(row.get("updated_at")) or _to_str(row.get("created_at"))
    return {
        "id": item_id,
        "asset_type": _to_str(row.get("asset_type")),
        "asset_status": _to_str(row.get("asset_status")),
        "key": _to_str(row.get("key")),
        "n_property_sn": _to_str(row.get("n_property_sn")),
        "property_sn": _to_str(row.get("property_sn")),
        "n_item_sn": _to_str(row.get("n_item_sn")),
        "item_sn": _to_str(row.get("item_sn")),
        "name": _to_str(row.get("name")),
        "name_code": _to_str(row.get("name_code")),
        "name_code2": _to_str(row.get("name_code2")),
        "model": _to_str(row.get("model")),
        "specification": _to_str(row.get("specification")),
        "unit": _to_str(row.get("unit")),
        "count": _to_int(row.get("count"), default=0),
        "purchase_date": _to_str(row.get("purchase_date")),
        "due_date": _to_str(row.get("due_date")),
        "return_date": _to_str(row.get("return_date")),
        "location": _to_str(row.get("location")),
        "memo": _to_str(row.get("memo")),
        "memo2": _to_str(row.get("memo2")),
        "keeper": _to_str(row.get("keeper")),
        "created_at": _to_str(row.get("created_at")),
        "created_by": _to_str(row.get("created_by")),
        "updated_at": _to_str(row.get("updated_at")),
        "updated_by": _to_str(row.get("updated_by")),
        "deleted_at": _to_str(row.get("deleted_at")),
        "deleted_by": _to_str(row.get("deleted_by")),
        "donated_at": donated_at,
        "donation_request_id": donation_request_id if donation_request_id > 0 else "",
    }


@contextmanager
def _locked_workbook():
    lock = FileLock(str(LOCK_PATH))
    with lock:
        wb = _load_workbook()
        yield wb


def _create_workbook() -> Workbook:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for sheet_name, headers in SHEETS.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)

    _seed_order_sn(wb["order_sn"])
    _seed_asset_status_codes(wb["asset_status_codes"])
    return wb


def _seed_order_sn(ws) -> bool:
    existing = {str(row[0]).strip() for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0]}
    added = False
    for name in ("asset", "item", "other"):
        if name not in existing:
            ws.append([name, 0])
            added = True
    return added


def _seed_asset_status_codes(ws) -> bool:
    existing_rows = _read_rows(ws)
    existing_codes = {_to_str(row.get("code")).strip() for row in existing_rows if not _is_blank(row.get("code"))}
    added = False
    now = _now_str()
    for code, description in DEFAULT_ASSET_STATUS_CODES:
        if code in existing_codes:
            continue
        ws.append([code, description, now, now])
        added = True
    return added


def _ensure_sheet(wb: Workbook, sheet_name: str, headers: list[str]) -> bool:
    changed = False
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        return True

    ws = wb[sheet_name]
    existing_headers = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
    if existing_headers != headers:
        rows = _read_rows(ws, existing_headers) if sheet_name != "inventory_items" else []
        ws.delete_rows(1, ws.max_row or 1)
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header, "") for header in headers])
        changed = True

    if sheet_name == "order_sn":
        if _seed_order_sn(ws):
            changed = True
    elif sheet_name == "asset_status_codes":
        if _seed_asset_status_codes(ws):
            changed = True

    return changed


def _normalize_string_fields(wb: Workbook, sheet_name: str, headers: list[str]) -> bool:
    string_fields = STRING_FIELDS.get(sheet_name)
    if not string_fields:
        return False

    ws = wb[sheet_name]
    rows = _read_rows(ws, headers)
    changed = False
    for row in rows:
        for field in string_fields:
            if row.get(field) is None:
                row[field] = ""
                changed = True

    if changed:
        _write_rows(ws, headers, rows)
    return changed


def _remove_removed_sheets(wb: Workbook) -> bool:
    changed = False
    for sheet_name in REMOVED_SHEETS:
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
            changed = True
    return changed


def _ensure_workbook(wb: Workbook) -> bool:
    changed = _remove_removed_sheets(wb)
    for sheet_name, headers in SHEETS.items():
        if _ensure_sheet(wb, sheet_name, headers):
            changed = True
        if _normalize_string_fields(wb, sheet_name, headers):
            changed = True
    return changed


def _load_workbook() -> Workbook:
    if not DB_PATH.exists():
        wb = _create_workbook()
        wb.save(DB_PATH)
        return wb

    wb = load_workbook(DB_PATH)
    if _ensure_workbook(wb):
        wb.save(DB_PATH)
    return wb


def _read_rows(ws, headers: list[str] | None = None) -> list[dict[str, Any]]:
    if headers is None:
        headers = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        values = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
        if all(_is_blank(value) for value in values.values()):
            continue
        rows.append(values)
    return rows


def _write_rows(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.delete_rows(1, ws.max_row or 1)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def _donation_map(
    donation_item_rows: list[dict[str, Any]],
    donation_request_rows: list[dict[str, Any]],
) -> dict[int, dict[str, Any]]:
    donation_date_map = {
        _to_int(row.get("id")): _to_str(row.get("created_at"))
        for row in donation_request_rows
    }
    item_map: dict[int, dict[str, Any]] = {}
    for row in donation_item_rows:
        request_id = _to_int(row.get("request_id"))
        item_id = _to_int(row.get("item_id"))
        if item_id <= 0:
            continue
        item_map[item_id] = {
            "donation_request_id": request_id,
            "donated_at": donation_date_map.get(request_id, ""),
        }
    return item_map


def _is_item_donated(
    row: dict[str, Any],
    *,
    donation_info: dict[str, Any] | None = None,
) -> bool:
    if _to_str(row.get("asset_status")) == "3":
        return True
    if donation_info and _to_int(donation_info.get("donation_request_id")) > 0:
        return True
    return False


def _next_id(rows: list[dict[str, Any]]) -> int:
    max_id = 0
    for row in rows:
        max_id = max(max_id, _to_int(row.get("id")))
    return max_id + 1


def init_db() -> None:
    _load_asset_category_mapping()
    with _locked_workbook() as wb:
        if _ensure_workbook(wb):
            wb.save(DB_PATH)


def _normalize_asset_status_code(value: Any) -> str:
    return _to_str(value).strip()


def _normalize_asset_status_description(value: Any) -> str:
    return _to_str(value).strip()


def _asset_status_sort_key(code: str) -> tuple[int, int | str]:
    stripped = code.strip()
    if stripped.isdigit():
        return (0, int(stripped))
    return (1, stripped)


def list_asset_status_codes() -> list[dict[str, Any]]:
    with _locked_workbook() as wb:
        rows = _read_rows(wb["asset_status_codes"])
    results = []
    for row in rows:
        code = _normalize_asset_status_code(row.get("code"))
        description = _normalize_asset_status_description(row.get("description"))
        if not code:
            continue
        results.append(
            {
                "code": code,
                "description": description,
            }
        )
    return sorted(results, key=lambda row: _asset_status_sort_key(row["code"]))


def create_asset_status_code(code: str, description: str) -> dict[str, Any]:
    normalized_code = _normalize_asset_status_code(code)
    normalized_description = _normalize_asset_status_description(description)
    if not normalized_code:
        raise ValueError("asset_status code is required")
    if not normalized_description:
        raise ValueError("asset_status description is required")

    with _locked_workbook() as wb:
        ws = wb["asset_status_codes"]
        rows = _read_rows(ws)
        existing_codes = {_normalize_asset_status_code(row.get("code")) for row in rows}
        if normalized_code in existing_codes:
            raise ValueError("asset_status code already exists")

        now = _now_str()
        rows.append(
            {
                "code": normalized_code,
                "description": normalized_description,
                "created_at": now,
                "updated_at": now,
            }
        )
        _write_rows(ws, SHEETS["asset_status_codes"], rows)
        wb.save(DB_PATH)
    return {"code": normalized_code, "description": normalized_description}


def update_asset_status_code(code: str, next_code: str, description: str) -> dict[str, Any]:
    normalized_code = _normalize_asset_status_code(code)
    normalized_next_code = _normalize_asset_status_code(next_code)
    normalized_description = _normalize_asset_status_description(description)
    if not normalized_code:
        raise ValueError("asset_status code is required")
    if not normalized_next_code:
        raise ValueError("asset_status code is required")
    if not normalized_description:
        raise ValueError("asset_status description is required")

    with _locked_workbook() as wb:
        code_ws = wb["asset_status_codes"]
        inventory_ws = wb["inventory_items"]
        code_rows = _read_rows(code_ws)
        inventory_rows = _read_rows(inventory_ws)

        target_row: dict[str, Any] | None = None
        for row in code_rows:
            if _normalize_asset_status_code(row.get("code")) == normalized_code:
                target_row = row
                break
        if target_row is None:
            raise ValueError("asset_status code not found")

        if normalized_next_code != normalized_code:
            for row in code_rows:
                if _normalize_asset_status_code(row.get("code")) == normalized_next_code:
                    raise ValueError("asset_status code already exists")

        now = _now_str()
        target_row["code"] = normalized_next_code
        target_row["description"] = normalized_description
        target_row["updated_at"] = now

        if normalized_next_code != normalized_code:
            for row in inventory_rows:
                if _is_blank(row.get("deleted_at")) and _normalize_asset_status_code(row.get("asset_status")) == normalized_code:
                    row["asset_status"] = normalized_next_code
                    row["updated_at"] = now
                    row["updated_by"] = "system"

        _write_rows(code_ws, SHEETS["asset_status_codes"], code_rows)
        _write_rows(inventory_ws, SHEETS["inventory_items"], inventory_rows)
        wb.save(DB_PATH)
    return {"code": normalized_next_code, "description": normalized_description}


def delete_asset_status_code(code: str) -> bool:
    normalized_code = _normalize_asset_status_code(code)
    if not normalized_code:
        raise ValueError("asset_status code is required")

    with _locked_workbook() as wb:
        code_ws = wb["asset_status_codes"]
        inventory_ws = wb["inventory_items"]
        code_rows = _read_rows(code_ws)
        inventory_rows = _read_rows(inventory_ws)

        remaining_rows = [row for row in code_rows if _normalize_asset_status_code(row.get("code")) != normalized_code]
        deleted = len(remaining_rows) != len(code_rows)
        if not deleted:
            raise ValueError("asset_status code not found")

        is_in_use = any(
            _is_blank(row.get("deleted_at")) and _normalize_asset_status_code(row.get("asset_status")) == normalized_code
            for row in inventory_rows
        )
        if is_in_use:
            raise ValueError("asset_status code is in use")

        _write_rows(code_ws, SHEETS["asset_status_codes"], remaining_rows)
        wb.save(DB_PATH)
    return True


def get_items_count() -> int:
    with _locked_workbook() as wb:
        rows = _read_rows(wb["inventory_items"])
    return sum(1 for row in rows if _is_blank(row.get("deleted_at")))


def get_pending_fix_count() -> int:
    with _locked_workbook() as wb:
        rows = _read_rows(wb["inventory_items"])
    return sum(
        1
        for row in rows
        if _is_blank(row.get("deleted_at"))
        and (
            _is_blank(_inventory_property_number(row))
            or _has_cjk(_inventory_property_number(row))
        )
    )


def _parse_request_date_value(value: Any) -> int:
    raw = _to_str(value).strip()
    if not raw:
        return 0
    for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
        try:
            return int(datetime.strptime(raw, fmt).timestamp())
        except ValueError:
            continue
    return 0


def get_dashboard_snapshot() -> dict[str, Any]:
    with _locked_workbook() as wb:
        inventory_rows = _read_rows(wb["inventory_items"])
        issue_request_rows = _read_rows(wb["issue_requests"])
        issue_item_rows = _read_rows(wb["issue_items"])
        borrow_request_rows = _read_rows(wb["borrow_requests"])
        borrow_item_rows = _read_rows(wb["borrow_items"])
        borrow_line_rows = _read_rows(wb["borrow_request_lines"])
        borrow_allocation_rows = _read_rows(wb["borrow_allocations"])
        donation_request_rows = _read_rows(wb["donation_requests"])
        donation_item_rows = _read_rows(wb["donation_items"])

    active_inventory_rows = [row for row in inventory_rows if _is_blank(row.get("deleted_at"))]
    donation_map = _donation_map(donation_item_rows, donation_request_rows)

    pending_fix_count = 0
    donated_count = 0
    category_counts: dict[str, int] = {}
    for row in active_inventory_rows:
        serial = _inventory_property_number(row)
        if _is_blank(serial) or _has_cjk(serial):
            pending_fix_count += 1
        donation_info = donation_map.get(_to_int(row.get("id")))
        if _is_item_donated(row, donation_info=donation_info):
            donated_count += 1
            continue
        if _to_str(row.get("asset_status")).strip() == "0":
            item_name = _to_str(row.get("name")).strip() or "未命名品項"
            category_counts[item_name] = category_counts.get(item_name, 0) + 1

    issue_item_count_map: dict[int, int] = {}
    for row in issue_item_rows:
        request_id = _to_int(row.get("request_id"))
        issue_item_count_map[request_id] = issue_item_count_map.get(request_id, 0) + 1

    borrow_item_count_map: dict[int, int] = {}
    for row in borrow_item_rows:
        request_id = _to_int(row.get("request_id"))
        borrow_item_count_map[request_id] = borrow_item_count_map.get(request_id, 0) + 1
    for row in borrow_line_rows:
        request_id = _to_int(row.get("request_id"))
        borrow_item_count_map[request_id] = borrow_item_count_map.get(request_id, 0) + 1

    borrow_request_allocations, _ = _build_borrow_allocation_stats(
        allocation_rows=borrow_allocation_rows,
        legacy_item_rows=borrow_item_rows,
    )
    borrow_requested_qty_by_request = _build_borrow_requested_qty_by_request(borrow_line_rows)

    donation_item_count_map: dict[int, int] = {}
    for row in donation_item_rows:
        request_id = _to_int(row.get("request_id"))
        donation_item_count_map[request_id] = donation_item_count_map.get(request_id, 0) + 1

    activities: list[dict[str, Any]] = []
    for row in issue_request_rows:
        request_id = _to_int(row.get("id"))
        item_count = issue_item_count_map.get(request_id, 0)
        activities.append(
            {
                "key": f"issue-{request_id}",
                "type": "領用",
                "dateLabel": _to_str(row.get("request_date")).strip() or "--",
                "dateValue": _parse_request_date_value(row.get("request_date")),
                "actor": _to_str(row.get("requester")).strip() or "未填寫",
                "summary": f"{item_count} 項品類" if item_count > 0 else "無品項資料",
                "requestId": str(request_id),
            }
        )

    reserved_borrow_count = 0
    overdue_borrow_count = 0
    due_soon_borrow_count = 0
    for row in borrow_request_rows:
        request_id = _to_int(row.get("id"))
        _normalize_borrow_status_in_place(
            row,
            has_allocations=borrow_request_allocations.get(request_id, 0) > 0,
            is_fully_allocated=borrow_request_allocations.get(request_id, 0) >= borrow_requested_qty_by_request.get(request_id, 0),
        )
        status = _to_str(row.get("status")).strip()
        item_count = borrow_item_count_map.get(request_id, 0)
        if status == "reserved":
            reserved_borrow_count += 1
        if status == "overdue":
            overdue_borrow_count += 1
        if status in {"borrowed", "overdue"} and _is_due_soon(
            due_date_value=row.get("due_date"),
            return_date_value=row.get("return_date"),
        ):
            due_soon_borrow_count += 1
        activities.append(
            {
                "key": f"borrow-{request_id}",
                "type": "借用",
                "dateLabel": _to_str(row.get("borrow_date")).strip() or "--",
                "dateValue": _parse_request_date_value(row.get("borrow_date")),
                "actor": _to_str(row.get("borrower")).strip() or "未填寫",
                "summary": f"{item_count} 項品類 · {status or '--'}" if item_count > 0 else f"無品項資料 · {status or '--'}",
                "requestId": str(request_id),
            }
        )

    for row in donation_request_rows:
        request_id = _to_int(row.get("id"))
        item_count = donation_item_count_map.get(request_id, 0)
        activities.append(
            {
                "key": f"donation-{request_id}",
                "type": "捐贈",
                "dateLabel": _to_str(row.get("donation_date")).strip() or "--",
                "dateValue": _parse_request_date_value(row.get("donation_date")),
                "actor": _to_str(row.get("donor")).strip() or "未填寫",
                "summary": f"{item_count} 項品類" if item_count > 0 else "無品項資料",
                "requestId": str(request_id),
            }
        )

    activities.sort(key=lambda row: (int(row.get("dateValue") or 0), _to_str(row.get("key"))), reverse=True)
    top_categories = sorted(category_counts.items(), key=lambda item: (-item[1], item[0]))
    return {
        "status": "success",
        "data": "這是管理系統的後端數據",
        "items": len(active_inventory_rows),
        "pendingFix": pending_fix_count,
        "totalRecords": len(issue_request_rows) + len(borrow_request_rows) + len(donation_request_rows),
        "reservedBorrowCount": reserved_borrow_count,
        "overdueBorrowCount": overdue_borrow_count,
        "dueSoonBorrowCount": due_soon_borrow_count,
        "donatedItemsCount": donated_count,
        "itemCategoryDistribution": [{"name": name, "count": count} for name, count in top_categories],
        "recentActivities": activities[:8],
    }


def list_items(*, include_donated: bool = False) -> list[dict[str, Any]]:
    with _locked_workbook() as wb:
        rows = _read_rows(wb["inventory_items"])
        donation_rows = _read_rows(wb["donation_items"])
        donation_request_rows = _read_rows(wb["donation_requests"])

    donation_map = _donation_map(donation_rows, donation_request_rows)
    results = []
    for row in rows:
        if not _is_blank(row.get("deleted_at")):
            continue
        donation_info = donation_map.get(_to_int(row.get("id")))
        if not include_donated and _is_item_donated(row, donation_info=donation_info):
            continue
        results.append(_to_inventory_api_row(row, donation_map=donation_map))
    return sorted(results, key=lambda row: _to_int(row.get("id")), reverse=True)


def get_item_by_id(item_id: int) -> dict[str, Any] | None:
    with _locked_workbook() as wb:
        rows = _read_rows(wb["inventory_items"])
        donation_rows = _read_rows(wb["donation_items"])
        donation_request_rows = _read_rows(wb["donation_requests"])
    donation_map = _donation_map(donation_rows, donation_request_rows)
    for row in rows:
        if _to_int(row.get("id")) == item_id and _is_blank(row.get("deleted_at")):
            return _to_inventory_api_row(row, donation_map=donation_map)
    return None


def create_item(item_data: dict[str, Any]) -> int:
    property_number = (
        _to_str(item_data.get("n_property_sn")).strip()
        or _to_str(item_data.get("property_sn")).strip()
        or _to_str(item_data.get("n_item_sn")).strip()
        or _to_str(item_data.get("item_sn")).strip()
    )
    if not property_number:
        order_sn_name = _asset_type_to_order_sn_name(item_data.get("asset_type"))
        order_sn_row = get_order_sn(order_sn_name)
        if order_sn_row is not None:
            property_number = order_sn_row["tmp_no"]

    with _locked_workbook() as wb:
        ws = wb["inventory_items"]
        rows = _read_rows(ws)
        new_id = _next_id(rows)
        rows.append(_to_inventory_create_row(new_id, item_data, property_number))
        _write_rows(ws, SHEETS["inventory_items"], rows)
        wb.save(DB_PATH)
        return new_id


def create_items_bulk(items: list[dict[str, Any]]) -> int:
    if not items:
        return 0

    with _locked_workbook() as wb:
        inventory_ws = wb["inventory_items"]
        order_ws = wb["order_sn"]
        inventory_rows = _read_rows(inventory_ws)
        order_rows = _read_rows(order_ws)

        order_map = {
            str(row.get("name", "")).strip(): row
            for row in order_rows
            if not _is_blank(row.get("name"))
        }
        for name in ("asset", "item", "other"):
            order_map.setdefault(name, {"name": name, "current_value": 0})

        next_id = _next_id(inventory_rows)
        created = 0

        for item_data in items:
            property_number = (
                _to_str(item_data.get("n_property_sn")).strip()
                or _to_str(item_data.get("property_sn")).strip()
                or _to_str(item_data.get("n_item_sn")).strip()
                or _to_str(item_data.get("item_sn")).strip()
            )
            if not property_number:
                order_sn_name = _asset_type_to_order_sn_name(item_data.get("asset_type"))
                order_row = order_map.get(order_sn_name)
                current_value = _to_int(order_row.get("current_value")) + 1
                order_row["current_value"] = current_value
                property_number = f"tmp-{_date_sn()}-{current_value:04d}"

            inventory_rows.append(_to_inventory_create_row(next_id, item_data, property_number))
            next_id += 1
            created += 1

        _write_rows(inventory_ws, SHEETS["inventory_items"], inventory_rows)
        _write_rows(order_ws, SHEETS["order_sn"], list(order_map.values()))
        wb.save(DB_PATH)
        return created


def update_item(item_id: int, item_data: dict[str, Any]) -> bool:
    with _locked_workbook() as wb:
        ws = wb["inventory_items"]
        rows = _read_rows(ws)
        updated = False
        for row in rows:
            if _to_int(row.get("id")) == item_id and _is_blank(row.get("deleted_at")):
                asset_type = _normalize_asset_type_input(item_data.get("asset_type"))
                n_property_sn = _to_str(item_data.get("n_property_sn")).strip()
                property_sn = _to_str(item_data.get("property_sn")).strip()
                n_item_sn = _to_str(item_data.get("n_item_sn")).strip()
                item_sn = _to_str(item_data.get("item_sn")).strip()
                property_number = n_property_sn or property_sn or n_item_sn or item_sn
                name_code, name_code2 = _normalize_name_codes(item_data.get("name_code"), item_data.get("name_code2"))
                count = 1
                next_key = _to_str(item_data.get("key")).strip()
                if not next_key:
                    tmp_row = {
                        "id": row.get("id"),
                        "key": row.get("key"),
                        "n_property_sn": n_property_sn,
                        "property_sn": property_sn,
                        "n_item_sn": n_item_sn,
                        "item_sn": item_sn,
                    }
                    next_key = _inventory_key(tmp_row, property_number)
                row.update(
                    {
                        "asset_type": asset_type,
                        "asset_status": _to_str(item_data.get("asset_status")).strip() or "0",
                        "key": next_key,
                        "n_property_sn": n_property_sn,
                        "property_sn": property_sn,
                        "n_item_sn": n_item_sn,
                        "item_sn": item_sn,
                        "name": _to_str(item_data.get("name")),
                        "name_code": name_code,
                        "name_code2": name_code2,
                        "model": _to_str(item_data.get("model")),
                        "specification": _to_str(item_data.get("specification")),
                        "unit": _to_str(item_data.get("unit")),
                        "count": str(count),
                        "purchase_date": _to_str(item_data.get("purchase_date")),
                        "due_date": _to_str(item_data.get("due_date")),
                        "return_date": _to_str(item_data.get("return_date")),
                        "location": _to_str(item_data.get("location")),
                        "keeper": _to_str(item_data.get("keeper")),
                        "memo": _to_str(item_data.get("memo")),
                        "memo2": _to_str(item_data.get("memo2")),
                        "updated_at": _now_str(),
                        "updated_by": "system",
                    }
                )
                updated = True
                break
        if updated:
            _write_rows(ws, SHEETS["inventory_items"], rows)
            wb.save(DB_PATH)
        return updated


def delete_item(item_id: int) -> bool:
    with _locked_workbook() as wb:
        ws = wb["inventory_items"]
        rows = _read_rows(ws)
        deleted = False
        for row in rows:
            if _to_int(row.get("id")) == item_id and _is_blank(row.get("deleted_at")):
                row["deleted_at"] = _now_str()
                row["deleted_by"] = "system"
                deleted = True
                break
        if deleted:
            _write_rows(ws, SHEETS["inventory_items"], rows)
            wb.save(DB_PATH)
        return deleted


def purge_soft_deleted_items() -> int:
    cutoff = datetime.now() - timedelta(days=180)
    with _locked_workbook() as wb:
        ws = wb["inventory_items"]
        rows = _read_rows(ws)
        kept: list[dict[str, Any]] = []
        deleted_count = 0
        for row in rows:
            deleted_at = row.get("deleted_at")
            if _is_blank(deleted_at):
                kept.append(row)
                continue
            try:
                deleted_time = datetime.strptime(str(deleted_at), "%Y-%m-%d %H:%M:%S")
            except ValueError:
                kept.append(row)
                continue
            if deleted_time <= cutoff:
                deleted_count += 1
            else:
                kept.append(row)
        if deleted_count:
            _write_rows(ws, SHEETS["inventory_items"], kept)
            wb.save(DB_PATH)
        return deleted_count


def log_inventory_action(
    action: str,
    *,
    entity: str = "inventory_item",
    entity_id: int | None = None,
    status: str = "success",
    detail: dict[str, Any] | None = None,
) -> None:
    serialized_detail = json.dumps(detail or {}, ensure_ascii=False)
    with _locked_workbook() as wb:
        ws = wb["operation_logs"]
        rows = _read_rows(ws)
        new_id = _next_id(rows)
        rows.append(
            {
                "id": new_id,
                "action": action,
                "entity": entity,
                "entity_id": entity_id if entity_id is not None else "",
                "status": status,
                "detail": serialized_detail,
                "created_at": _now_str(),
            }
        )
        _write_rows(ws, SHEETS["operation_logs"], rows)
        wb.save(DB_PATH)


def _append_operation_log_entry(
    operation_rows: list[dict[str, Any]],
    *,
    action: str,
    entity: str = "inventory_item",
    entity_id: int | None = None,
    status: str = "success",
    detail: dict[str, Any] | None = None,
) -> None:
    operation_rows.append(
        {
            "id": _next_id(operation_rows),
            "action": action,
            "entity": entity,
            "entity_id": entity_id if entity_id is not None else "",
            "status": status,
            "detail": json.dumps(detail or {}, ensure_ascii=False),
            "created_at": _now_str(),
        }
    )


def _append_movement_ledger_entry(
    movement_rows: list[dict[str, Any]],
    *,
    item_id: int,
    from_status: str,
    to_status: str,
    action: str,
    entity: str,
    entity_id: int | None,
    operator: str = "system",
) -> None:
    movement_rows.append(
        {
            "id": _next_id(movement_rows),
            "item_id": item_id,
            "from_status": from_status,
            "to_status": to_status,
            "action": action,
            "entity": entity,
            "entity_id": entity_id if entity_id is not None else "",
            "operator": operator,
            "created_at": _now_str(),
        }
    )


def _set_inventory_status_with_movement(
    *,
    movement_rows: list[dict[str, Any]],
    row: dict[str, Any],
    item_id: int,
    status: str,
    action: str,
    entity: str,
    entity_id: int | None,
    operator: str = "system",
) -> None:
    previous_status = _to_str(row.get("asset_status")).strip()
    _set_inventory_status(row, status)
    _append_movement_ledger_entry(
        movement_rows,
        item_id=item_id,
        from_status=previous_status,
        to_status=status,
        action=action,
        entity=entity,
        entity_id=entity_id,
        operator=operator,
    )


def _parse_log_time(value: Any) -> datetime | None:
    raw = _to_str(value).strip()
    if not raw:
        return None
    try:
        return datetime.strptime(raw, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None


def _matches_time_filter(*, created_at: Any, start_at: datetime | None, end_at: datetime | None) -> bool:
    if start_at is None and end_at is None:
        return True
    created_time = _parse_log_time(created_at)
    if created_time is None:
        return False
    if start_at is not None and created_time < start_at:
        return False
    if end_at is not None and created_time > end_at:
        return False
    return True


def _parse_json_detail(value: Any) -> dict[str, Any]:
    raw = _to_str(value).strip()
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _matches_item_filter_from_detail(detail: dict[str, Any], item_id: int) -> bool:
    detail_item_id = _to_int(detail.get("item_id"))
    if detail_item_id == item_id:
        return True
    item_ids = detail.get("item_ids")
    if isinstance(item_ids, list):
        return item_id in {_to_int(value) for value in item_ids}
    return False


def _month_key_from_datetime(value: datetime) -> str:
    return value.strftime("%Y%m")


def _archive_workbook_path(month_key: str) -> Path:
    return LOG_ARCHIVE_DIR / f"logs_{month_key}.xlsx"


def _ensure_archive_dir() -> None:
    LOG_ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)


def _list_archive_month_keys() -> list[str]:
    if not LOG_ARCHIVE_DIR.exists():
        return []
    keys: list[str] = []
    for path in LOG_ARCHIVE_DIR.iterdir():
        match = ARCHIVE_FILE_PATTERN.match(path.name)
        if match:
            keys.append(match.group(1))
    return sorted(set(keys))


def _month_bounds(month_key: str) -> tuple[datetime, datetime] | None:
    if len(month_key) != 6 or not month_key.isdigit():
        return None
    year = int(month_key[:4])
    month = int(month_key[4:6])
    if month < 1 or month > 12:
        return None
    start = datetime(year, month, 1)
    if month == 12:
        next_month = datetime(year + 1, 1, 1)
    else:
        next_month = datetime(year, month + 1, 1)
    end = next_month - timedelta(seconds=1)
    return start, end


def _month_intersects_range(month_key: str, start_at: datetime | None, end_at: datetime | None) -> bool:
    bounds = _month_bounds(month_key)
    if bounds is None:
        return False
    month_start, month_end = bounds
    if start_at is not None and month_end < start_at:
        return False
    if end_at is not None and month_start > end_at:
        return False
    return True


def _candidate_archive_month_keys(start_at: datetime | None, end_at: datetime | None) -> list[str]:
    keys = _list_archive_month_keys()
    if start_at is None and end_at is None:
        return keys
    return [key for key in keys if _month_intersects_range(key, start_at, end_at)]


def _load_archive_workbook(path: Path) -> Workbook:
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
    _ensure_sheet(wb, "operation_logs", SHEETS["operation_logs"])
    _ensure_sheet(wb, "movement_ledger", SHEETS["movement_ledger"])
    return wb


def _append_rows_to_archive(sheet_name: str, rows: list[dict[str, Any]], month_key: str) -> None:
    if not rows:
        return
    _ensure_archive_dir()
    archive_path = _archive_workbook_path(month_key)
    lock_path = archive_path.with_suffix(".lock")
    lock = FileLock(str(lock_path))
    with lock:
        wb = _load_archive_workbook(archive_path)
        ws = wb[sheet_name]
        headers = SHEETS[sheet_name]
        for row in rows:
            ws.append([row.get(header, "") for header in headers])
        wb.save(archive_path)


def archive_old_logs(*, retention_days: int = HOT_LOG_RETENTION_DAYS) -> dict[str, int]:
    cutoff = datetime.now() - timedelta(days=retention_days)
    with _locked_workbook() as wb:
        operation_rows = _read_rows(wb["operation_logs"])
        movement_rows = _read_rows(wb["movement_ledger"])

        kept_operations: list[dict[str, Any]] = []
        kept_movements: list[dict[str, Any]] = []
        archived_operations_by_month: dict[str, list[dict[str, Any]]] = {}
        archived_movements_by_month: dict[str, list[dict[str, Any]]] = {}

        for row in operation_rows:
            created_time = _parse_log_time(row.get("created_at"))
            if created_time is None or created_time >= cutoff:
                kept_operations.append(row)
                continue
            month_key = _month_key_from_datetime(created_time)
            archived_operations_by_month.setdefault(month_key, []).append(row)

        for row in movement_rows:
            created_time = _parse_log_time(row.get("created_at"))
            if created_time is None or created_time >= cutoff:
                kept_movements.append(row)
                continue
            month_key = _month_key_from_datetime(created_time)
            archived_movements_by_month.setdefault(month_key, []).append(row)

        operation_archived = len(operation_rows) - len(kept_operations)
        movement_archived = len(movement_rows) - len(kept_movements)
        if operation_archived:
            _write_rows(wb["operation_logs"], SHEETS["operation_logs"], kept_operations)
        if movement_archived:
            _write_rows(wb["movement_ledger"], SHEETS["movement_ledger"], kept_movements)
        if operation_archived or movement_archived:
            wb.save(DB_PATH)

    for month_key, rows in archived_operations_by_month.items():
        _append_rows_to_archive("operation_logs", rows, month_key)
    for month_key, rows in archived_movements_by_month.items():
        _append_rows_to_archive("movement_ledger", rows, month_key)

    return {
        "operation_logs_archived": operation_archived,
        "movement_ledger_archived": movement_archived,
    }


def _list_archive_sheet_rows(sheet_name: str, *, start_at: datetime | None, end_at: datetime | None) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for month_key in _candidate_archive_month_keys(start_at, end_at):
        archive_path = _archive_workbook_path(month_key)
        if not archive_path.exists():
            continue
        lock = FileLock(str(archive_path.with_suffix(".lock")))
        with lock:
            wb = _load_archive_workbook(archive_path)
            if sheet_name not in wb.sheetnames:
                continue
            rows.extend(_read_rows(wb[sheet_name]))
    return rows


def list_movement_ledger(
    *,
    start_at: datetime | None = None,
    end_at: datetime | None = None,
    action: str = "",
    entity: str = "",
    item_id: int | None = None,
    entity_id: int | None = None,
    scope: str = "hot",
) -> list[dict[str, Any]]:
    with _locked_workbook() as wb:
        movement_rows = _read_rows(wb["movement_ledger"])
        inventory_rows = _read_rows(wb["inventory_items"])
    if scope == "all":
        movement_rows.extend(_list_archive_sheet_rows("movement_ledger", start_at=start_at, end_at=end_at))

    inventory_map = {
        _to_int(row.get("id")): {
            "item_name": _to_str(row.get("name")),
            "item_model": _to_str(row.get("model")),
        }
        for row in inventory_rows
    }
    normalized_action = _to_str(action).strip().lower()
    normalized_entity = _to_str(entity).strip().lower()
    filtered: list[dict[str, Any]] = []
    for row in movement_rows:
        row_action = _to_str(row.get("action")).strip()
        row_entity = _to_str(row.get("entity")).strip()
        row_item_id = _to_int(row.get("item_id"))
        row_entity_id = _to_int(row.get("entity_id"))
        if normalized_action and row_action.lower() != normalized_action:
            continue
        if normalized_entity and row_entity.lower() != normalized_entity:
            continue
        if item_id is not None and row_item_id != item_id:
            continue
        if entity_id is not None and row_entity_id != entity_id:
            continue
        if not _matches_time_filter(created_at=row.get("created_at"), start_at=start_at, end_at=end_at):
            continue
        item_info = inventory_map.get(row_item_id, {})
        filtered.append(
            {
                "id": _to_int(row.get("id")),
                "item_id": row_item_id,
                "item_name": _to_str(item_info.get("item_name")),
                "item_model": _to_str(item_info.get("item_model")),
                "from_status": _to_str(row.get("from_status")),
                "to_status": _to_str(row.get("to_status")),
                "action": row_action,
                "entity": row_entity,
                "entity_id": row_entity_id if row_entity_id > 0 else None,
                "operator": _to_str(row.get("operator")),
                "created_at": _to_str(row.get("created_at")),
            }
        )

    return sorted(filtered, key=lambda entry: _to_int(entry.get("id")), reverse=True)


def list_operation_logs(
    *,
    start_at: datetime | None = None,
    end_at: datetime | None = None,
    action: str = "",
    entity: str = "",
    item_id: int | None = None,
    entity_id: int | None = None,
    scope: str = "hot",
) -> list[dict[str, Any]]:
    with _locked_workbook() as wb:
        rows = _read_rows(wb["operation_logs"])
    if scope == "all":
        rows.extend(_list_archive_sheet_rows("operation_logs", start_at=start_at, end_at=end_at))

    normalized_action = _to_str(action).strip().lower()
    normalized_entity = _to_str(entity).strip().lower()
    filtered: list[dict[str, Any]] = []
    for row in rows:
        row_action = _to_str(row.get("action")).strip()
        row_entity = _to_str(row.get("entity")).strip()
        row_entity_id = _to_int(row.get("entity_id"))
        detail: dict[str, Any] = {}
        if normalized_action and row_action.lower() != normalized_action:
            continue
        if normalized_entity and row_entity.lower() != normalized_entity:
            continue
        if entity_id is not None and row_entity_id != entity_id:
            continue
        if item_id is not None:
            detail = _parse_json_detail(row.get("detail"))
            if not _matches_item_filter_from_detail(detail, item_id):
                continue
        if not _matches_time_filter(created_at=row.get("created_at"), start_at=start_at, end_at=end_at):
            continue
        if not detail:
            detail = _parse_json_detail(row.get("detail"))
        filtered.append(
            {
                "id": _to_int(row.get("id")),
                "action": row_action,
                "entity": row_entity,
                "entity_id": row_entity_id if row_entity_id > 0 else None,
                "status": _to_str(row.get("status")),
                "detail": detail,
                "created_at": _to_str(row.get("created_at")),
            }
        )

    return sorted(filtered, key=lambda entry: _to_int(entry.get("id")), reverse=True)


def get_order_sn(name: str) -> dict[str, Any] | None:
    with _locked_workbook() as wb:
        ws = wb["order_sn"]
        rows = _read_rows(ws)
        for row in rows:
            if str(row.get("name", "")).strip() == name:
                current_value = _to_int(row.get("current_value")) + 1
                row["current_value"] = current_value
                tmp_no = f"tmp-{_date_sn()}-{current_value:04d}"
                _write_rows(ws, SHEETS["order_sn"], rows)
                wb.save(DB_PATH)
                return {"tmp_no": tmp_no}
        return None


def validate_item_ids_available(
    item_ids: list[int],
    *,
    allow_donation_request_id: int | None = None,
    enforce_unique: bool = False,
) -> tuple[bool, str | None]:
    with _locked_workbook() as wb:
        inventory_rows = _read_rows(wb["inventory_items"])
        donation_item_rows = _read_rows(wb["donation_items"])

    if enforce_unique and len(item_ids) != len(set(item_ids)):
        return False, "item_id cannot be duplicated"

    inventory_map = {_to_int(row.get("id")): row for row in inventory_rows if _is_blank(row.get("deleted_at"))}
    donation_request_map = {
        _to_int(row.get("item_id")): _to_int(row.get("request_id"))
        for row in donation_item_rows
    }
    for item_id in item_ids:
        row = inventory_map.get(item_id)
        if row is None:
            return False, f"item_id {item_id} not found"
        donation_request_id = donation_request_map.get(item_id, 0)
        if _to_str(row.get("asset_status")) != "3" and donation_request_id == 0:
            continue
        if allow_donation_request_id is not None and donation_request_id == allow_donation_request_id:
            continue
        return False, f"item_id {item_id} is already donated"

    return True, None


def _normalize_request_item_ids(items: list[dict[str, Any]]) -> list[int]:
    item_ids: list[int] = []
    seen: set[int] = set()
    for item in items:
        item_id = _to_int(item.get("item_id"))
        quantity = _to_int(item.get("quantity"))
        if item_id <= 0:
            raise ValueError(f"item_id {item_id} not found")
        if quantity != 1:
            raise ValueError("quantity must be 1 in single-item mode")
        if item_id in seen:
            raise ValueError("item_id cannot be duplicated")
        seen.add(item_id)
        item_ids.append(item_id)
    return item_ids


def _active_inventory_rows_map(inventory_rows: list[dict[str, Any]]) -> dict[int, dict[str, Any]]:
    return {
        _to_int(row.get("id")): row
        for row in inventory_rows
        if _is_blank(row.get("deleted_at"))
    }


def _set_inventory_status(row: dict[str, Any], status: str) -> None:
    row["asset_status"] = status
    row["updated_at"] = _now_str()
    row["updated_by"] = "system"


def _validate_item_status(
    *,
    inventory_map: dict[int, dict[str, Any]],
    item_id: int,
    allowed_statuses: set[str],
) -> dict[str, Any]:
    row = inventory_map.get(item_id)
    if row is None:
        raise ValueError(f"item_id {item_id} not found")
    status = _to_str(row.get("asset_status")).strip()
    if status not in allowed_statuses:
        raise ValueError(f"item_id {item_id} is unavailable")
    return row

def create_issue_request(request_data: dict[str, Any], items: list[dict[str, Any]]) -> int:
    with _locked_workbook() as wb:
        request_ws = wb["issue_requests"]
        item_ws = wb["issue_items"]
        inventory_ws = wb["inventory_items"]
        movement_ws = wb["movement_ledger"]
        request_rows = _read_rows(request_ws)
        item_rows = _read_rows(item_ws)
        inventory_rows = _read_rows(inventory_ws)
        movement_rows = _read_rows(movement_ws)
        selected_item_ids = _normalize_request_item_ids(items)
        inventory_map = _active_inventory_rows_map(inventory_rows)
        request_id = _next_id(request_rows)
        for item_id in selected_item_ids:
            row = _validate_item_status(inventory_map=inventory_map, item_id=item_id, allowed_statuses={"0"})
            _set_inventory_status_with_movement(
                movement_rows=movement_rows,
                row=row,
                item_id=item_id,
                status="1",
                action="create",
                entity="issue_request",
                entity_id=request_id,
            )
        request_rows.append(
            {
                "id": request_id,
                "requester": request_data["requester"],
                "department": request_data["department"],
                "purpose": request_data["purpose"],
                "request_date": request_data["request_date"],
                "memo": request_data["memo"],
                "created_at": _now_str(),
            }
        )
        next_item_id = _next_id(item_rows)
        for index, item in enumerate(items):
            item_rows.append(
                {
                    "id": next_item_id + index,
                    "request_id": request_id,
                    "item_id": item["item_id"],
                    "quantity": 1,
                    "note": item.get("note", ""),
                }
            )
        _write_rows(request_ws, SHEETS["issue_requests"], request_rows)
        _write_rows(item_ws, SHEETS["issue_items"], item_rows)
        _write_rows(inventory_ws, SHEETS["inventory_items"], inventory_rows)
        _write_rows(movement_ws, SHEETS["movement_ledger"], movement_rows)
        wb.save(DB_PATH)
        return request_id


def list_issue_requests() -> list[dict[str, Any]]:
    with _locked_workbook() as wb:
        rows = _read_rows(wb["issue_requests"])
    return sorted(rows, key=lambda row: _to_int(row.get("id")), reverse=True)


def get_issue_request(request_id: int) -> dict[str, Any] | None:
    with _locked_workbook() as wb:
        rows = _read_rows(wb["issue_requests"])
    for row in rows:
        if _to_int(row.get("id")) == request_id:
            return row
    return None


def _active_inventory_detail_map(inventory_rows: list[dict[str, Any]]) -> dict[int, dict[str, Any]]:
    return {
        _to_int(row.get("id")): {
            "item_name": row.get("name", ""),
            "item_model": row.get("model", ""),
        }
        for row in inventory_rows
        if _is_blank(row.get("deleted_at"))
    }


def list_issue_items_map(request_ids: set[int] | None = None) -> dict[int, list[dict[str, Any]]]:
    with _locked_workbook() as wb:
        item_rows = _read_rows(wb["issue_items"])
        inventory_rows = _read_rows(wb["inventory_items"])

    inventory_map = _active_inventory_detail_map(inventory_rows)
    selected_ids = request_ids or set()
    results: dict[int, list[dict[str, Any]]] = {}
    for row in item_rows:
        request_id = _to_int(row.get("request_id"))
        if selected_ids and request_id not in selected_ids:
            continue
        item_id = _to_int(row.get("item_id"))
        details = inventory_map.get(item_id, {"item_name": None, "item_model": None})
        results.setdefault(request_id, []).append(
            {
                "id": row.get("id"),
                "request_id": row.get("request_id"),
                "item_id": row.get("item_id"),
                "quantity": row.get("quantity"),
                "note": row.get("note", ""),
                "item_name": details.get("item_name"),
                "item_model": details.get("item_model"),
            }
        )

    for grouped_rows in results.values():
        grouped_rows.sort(key=lambda row: _to_int(row.get("id")))
    return results


def list_issue_items(request_id: int) -> list[dict[str, Any]]:
    return list_issue_items_map({request_id}).get(request_id, [])


def update_issue_request(request_id: int, request_data: dict[str, Any], items: list[dict[str, Any]]) -> bool:
    with _locked_workbook() as wb:
        request_ws = wb["issue_requests"]
        item_ws = wb["issue_items"]
        inventory_ws = wb["inventory_items"]
        movement_ws = wb["movement_ledger"]
        request_rows = _read_rows(request_ws)
        item_rows = _read_rows(item_ws)
        inventory_rows = _read_rows(inventory_ws)
        movement_rows = _read_rows(movement_ws)
        updated = False
        for row in request_rows:
            if _to_int(row.get("id")) == request_id:
                row.update(
                    {
                        "requester": request_data["requester"],
                        "department": request_data["department"],
                        "purpose": request_data["purpose"],
                        "request_date": request_data["request_date"],
                        "memo": request_data["memo"],
                    }
                )
                updated = True
                break
        if not updated:
            return False

        old_items = [row for row in item_rows if _to_int(row.get("request_id")) == request_id]
        old_item_ids = {_to_int(row.get("item_id")) for row in old_items}
        new_item_ids = set(_normalize_request_item_ids(items))
        inventory_map = _active_inventory_rows_map(inventory_rows)

        for item_id in new_item_ids - old_item_ids:
            row = _validate_item_status(inventory_map=inventory_map, item_id=item_id, allowed_statuses={"0"})
            _set_inventory_status_with_movement(
                movement_rows=movement_rows,
                row=row,
                item_id=item_id,
                status="1",
                action="update",
                entity="issue_request",
                entity_id=request_id,
            )
        for item_id in new_item_ids & old_item_ids:
            row = _validate_item_status(inventory_map=inventory_map, item_id=item_id, allowed_statuses={"0", "1"})
            _set_inventory_status_with_movement(
                movement_rows=movement_rows,
                row=row,
                item_id=item_id,
                status="1",
                action="update",
                entity="issue_request",
                entity_id=request_id,
            )
        for item_id in old_item_ids - new_item_ids:
            row = _validate_item_status(inventory_map=inventory_map, item_id=item_id, allowed_statuses={"1"})
            _set_inventory_status_with_movement(
                movement_rows=movement_rows,
                row=row,
                item_id=item_id,
                status="0",
                action="update",
                entity="issue_request",
                entity_id=request_id,
            )

        item_rows = [row for row in item_rows if _to_int(row.get("request_id")) != request_id]
        next_item_id = _next_id(item_rows)
        for index, item in enumerate(items):
            item_rows.append(
                {
                    "id": next_item_id + index,
                    "request_id": request_id,
                    "item_id": item["item_id"],
                    "quantity": 1,
                    "note": item.get("note", ""),
                }
            )
        _write_rows(request_ws, SHEETS["issue_requests"], request_rows)
        _write_rows(item_ws, SHEETS["issue_items"], item_rows)
        _write_rows(inventory_ws, SHEETS["inventory_items"], inventory_rows)
        _write_rows(movement_ws, SHEETS["movement_ledger"], movement_rows)
        wb.save(DB_PATH)
        return True


def delete_issue_request(request_id: int) -> bool:
    with _locked_workbook() as wb:
        request_ws = wb["issue_requests"]
        item_ws = wb["issue_items"]
        inventory_ws = wb["inventory_items"]
        movement_ws = wb["movement_ledger"]
        request_rows = _read_rows(request_ws)
        item_rows = _read_rows(item_ws)
        inventory_rows = _read_rows(inventory_ws)
        movement_rows = _read_rows(movement_ws)
        remaining_requests = [row for row in request_rows if _to_int(row.get("id")) != request_id]
        deleted = len(remaining_requests) != len(request_rows)
        if not deleted:
            return False
        old_items = [row for row in item_rows if _to_int(row.get("request_id")) == request_id]
        inventory_map = _active_inventory_rows_map(inventory_rows)
        for old_item in old_items:
            item_id = _to_int(old_item.get("item_id"))
            row = _validate_item_status(inventory_map=inventory_map, item_id=item_id, allowed_statuses={"1"})
            _set_inventory_status_with_movement(
                movement_rows=movement_rows,
                row=row,
                item_id=item_id,
                status="0",
                action="delete",
                entity="issue_request",
                entity_id=request_id,
            )
        remaining_items = [row for row in item_rows if _to_int(row.get("request_id")) != request_id]
        _write_rows(request_ws, SHEETS["issue_requests"], remaining_requests)
        _write_rows(item_ws, SHEETS["issue_items"], remaining_items)
        _write_rows(inventory_ws, SHEETS["inventory_items"], inventory_rows)
        _write_rows(movement_ws, SHEETS["movement_ledger"], movement_rows)
        wb.save(DB_PATH)
        return True


def create_donation_request(request_data: dict[str, Any], items: list[dict[str, Any]]) -> int:
    with _locked_workbook() as wb:
        request_ws = wb["donation_requests"]
        item_ws = wb["donation_items"]
        inventory_ws = wb["inventory_items"]
        movement_ws = wb["movement_ledger"]
        request_rows = _read_rows(request_ws)
        item_rows = _read_rows(item_ws)
        inventory_rows = _read_rows(inventory_ws)
        movement_rows = _read_rows(movement_ws)

        selected_item_ids = _normalize_request_item_ids(items)
        inventory_map = _active_inventory_rows_map(inventory_rows)
        request_id = _next_id(request_rows)
        for item_id in selected_item_ids:
            row = _validate_item_status(inventory_map=inventory_map, item_id=item_id, allowed_statuses={"0"})
            _set_inventory_status_with_movement(
                movement_rows=movement_rows,
                row=row,
                item_id=item_id,
                status="3",
                action="create",
                entity="donation_request",
                entity_id=request_id,
            )
        request_rows.append(
            {
                "id": request_id,
                "donor": request_data["donor"],
                "department": request_data["department"],
                "recipient": request_data["recipient"],
                "purpose": request_data["purpose"],
                "donation_date": request_data["donation_date"],
                "memo": request_data["memo"],
                "created_at": _now_str(),
            }
        )

        next_item_id = _next_id(item_rows)
        for index, item in enumerate(items):
            item_rows.append(
                {
                    "id": next_item_id + index,
                    "request_id": request_id,
                    "item_id": item["item_id"],
                    "quantity": 1,
                    "note": item.get("note", ""),
                }
            )

        _write_rows(request_ws, SHEETS["donation_requests"], request_rows)
        _write_rows(item_ws, SHEETS["donation_items"], item_rows)
        _write_rows(inventory_ws, SHEETS["inventory_items"], inventory_rows)
        _write_rows(movement_ws, SHEETS["movement_ledger"], movement_rows)
        wb.save(DB_PATH)
        return request_id


def list_donation_requests() -> list[dict[str, Any]]:
    with _locked_workbook() as wb:
        rows = _read_rows(wb["donation_requests"])
    return sorted(rows, key=lambda row: _to_int(row.get("id")), reverse=True)


def get_donation_request(request_id: int) -> dict[str, Any] | None:
    with _locked_workbook() as wb:
        rows = _read_rows(wb["donation_requests"])
    for row in rows:
        if _to_int(row.get("id")) == request_id:
            return row
    return None


def list_donation_items_map(request_ids: set[int] | None = None) -> dict[int, list[dict[str, Any]]]:
    with _locked_workbook() as wb:
        item_rows = _read_rows(wb["donation_items"])
        inventory_rows = _read_rows(wb["inventory_items"])

    inventory_map = _active_inventory_detail_map(inventory_rows)
    selected_ids = request_ids or set()
    results: dict[int, list[dict[str, Any]]] = {}
    for row in item_rows:
        request_id = _to_int(row.get("request_id"))
        if selected_ids and request_id not in selected_ids:
            continue
        item_id = _to_int(row.get("item_id"))
        details = inventory_map.get(item_id, {"item_name": None, "item_model": None})
        results.setdefault(request_id, []).append(
            {
                "id": row.get("id"),
                "request_id": row.get("request_id"),
                "item_id": row.get("item_id"),
                "quantity": row.get("quantity"),
                "note": row.get("note", ""),
                "item_name": details.get("item_name"),
                "item_model": details.get("item_model"),
            }
        )

    for grouped_rows in results.values():
        grouped_rows.sort(key=lambda row: _to_int(row.get("id")))
    return results


def list_donation_items(request_id: int) -> list[dict[str, Any]]:
    return list_donation_items_map({request_id}).get(request_id, [])


def update_donation_request(request_id: int, request_data: dict[str, Any], items: list[dict[str, Any]]) -> bool:
    with _locked_workbook() as wb:
        request_ws = wb["donation_requests"]
        item_ws = wb["donation_items"]
        inventory_ws = wb["inventory_items"]
        movement_ws = wb["movement_ledger"]
        request_rows = _read_rows(request_ws)
        item_rows = _read_rows(item_ws)
        inventory_rows = _read_rows(inventory_ws)
        movement_rows = _read_rows(movement_ws)

        request_row = None
        for row in request_rows:
            if _to_int(row.get("id")) == request_id:
                request_row = row
                break
        if request_row is None:
            return False

        new_item_ids = set(_normalize_request_item_ids(items))
        old_items = [row for row in item_rows if _to_int(row.get("request_id")) == request_id]
        old_item_ids = {_to_int(row.get("item_id")) for row in old_items}

        inventory_map = _active_inventory_rows_map(inventory_rows)
        for item_id in new_item_ids - old_item_ids:
            row = _validate_item_status(inventory_map=inventory_map, item_id=item_id, allowed_statuses={"0"})
            _set_inventory_status_with_movement(
                movement_rows=movement_rows,
                row=row,
                item_id=item_id,
                status="3",
                action="update",
                entity="donation_request",
                entity_id=request_id,
            )
        for item_id in new_item_ids & old_item_ids:
            row = _validate_item_status(inventory_map=inventory_map, item_id=item_id, allowed_statuses={"0", "3"})
            _set_inventory_status_with_movement(
                movement_rows=movement_rows,
                row=row,
                item_id=item_id,
                status="3",
                action="update",
                entity="donation_request",
                entity_id=request_id,
            )
        for item_id in old_item_ids - new_item_ids:
            row = _validate_item_status(inventory_map=inventory_map, item_id=item_id, allowed_statuses={"3"})
            _set_inventory_status_with_movement(
                movement_rows=movement_rows,
                row=row,
                item_id=item_id,
                status="0",
                action="update",
                entity="donation_request",
                entity_id=request_id,
            )

        request_row.update(
            {
                "donor": request_data["donor"],
                "department": request_data["department"],
                "recipient": request_data["recipient"],
                "purpose": request_data["purpose"],
                "donation_date": request_data["donation_date"],
                "memo": request_data["memo"],
            }
        )

        item_rows = [row for row in item_rows if _to_int(row.get("request_id")) != request_id]
        next_item_id = _next_id(item_rows)
        for index, item in enumerate(items):
            item_rows.append(
                {
                    "id": next_item_id + index,
                    "request_id": request_id,
                    "item_id": item["item_id"],
                    "quantity": 1,
                    "note": item.get("note", ""),
                }
            )

        _write_rows(request_ws, SHEETS["donation_requests"], request_rows)
        _write_rows(item_ws, SHEETS["donation_items"], item_rows)
        _write_rows(inventory_ws, SHEETS["inventory_items"], inventory_rows)
        _write_rows(movement_ws, SHEETS["movement_ledger"], movement_rows)
        wb.save(DB_PATH)
        return True


def delete_donation_request(request_id: int) -> bool:
    with _locked_workbook() as wb:
        request_ws = wb["donation_requests"]
        item_ws = wb["donation_items"]
        inventory_ws = wb["inventory_items"]
        movement_ws = wb["movement_ledger"]
        request_rows = _read_rows(request_ws)
        item_rows = _read_rows(item_ws)
        inventory_rows = _read_rows(inventory_ws)
        movement_rows = _read_rows(movement_ws)

        remaining_requests = [row for row in request_rows if _to_int(row.get("id")) != request_id]
        deleted = len(remaining_requests) != len(request_rows)
        if not deleted:
            return False

        removed_items = [row for row in item_rows if _to_int(row.get("request_id")) == request_id]
        remaining_items = [row for row in item_rows if _to_int(row.get("request_id")) != request_id]
        inventory_map = _active_inventory_rows_map(inventory_rows)
        for removed_item in removed_items:
            item_id = _to_int(removed_item.get("item_id"))
            row = _validate_item_status(inventory_map=inventory_map, item_id=item_id, allowed_statuses={"3"})
            _set_inventory_status_with_movement(
                movement_rows=movement_rows,
                row=row,
                item_id=item_id,
                status="0",
                action="delete",
                entity="donation_request",
                entity_id=request_id,
            )

        _write_rows(request_ws, SHEETS["donation_requests"], remaining_requests)
        _write_rows(item_ws, SHEETS["donation_items"], remaining_items)
        _write_rows(inventory_ws, SHEETS["inventory_items"], inventory_rows)
        _write_rows(movement_ws, SHEETS["movement_ledger"], movement_rows)
        wb.save(DB_PATH)
        return True


def _borrow_status_uses_inventory(status: Any) -> bool:
    normalized = _to_str(status).strip().lower()
    return normalized in {"partial_borrowed", "borrowed", "overdue"}


def _parse_request_date(value: Any) -> date | None:
    raw = _to_str(value).strip()
    if not raw:
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _borrow_key(item_name: Any, item_model: Any) -> tuple[str, str]:
    return (_to_str(item_name).strip(), _to_str(item_model).strip())


def _derive_borrow_status(
    *,
    due_date_value: Any,
    return_date_value: Any,
    status_value: Any = "",
    borrow_date_value: Any = "",
    has_allocations: bool = False,
    is_fully_allocated: bool = True,
    today: date | None = None,
) -> str:
    now = today or date.today()
    return_date = _parse_request_date(return_date_value)
    if return_date is not None:
        return "returned"

    status = _to_str(status_value).strip().lower()
    borrow_date = _parse_request_date(borrow_date_value)
    if not has_allocations:
        if borrow_date is not None:
            overdue_days = (now - borrow_date).days
            if overdue_days >= BORROW_RESERVATION_CANCEL_AFTER_DAYS:
                return "cancelled"
            if overdue_days >= 1:
                return "expired"
        if status in {"expired", "cancelled"}:
            return status
        return "reserved"

    due_date = _parse_request_date(due_date_value)
    if due_date is not None and due_date < now:
        return "overdue"
    if not is_fully_allocated:
        return "partial_borrowed"
    return "borrowed"


def _is_due_soon(*, due_date_value: Any, return_date_value: Any, today: date | None = None, days: int = 3) -> bool:
    now = today or date.today()
    if _parse_request_date(return_date_value) is not None:
        return False
    due_date = _parse_request_date(due_date_value)
    if due_date is None:
        return False
    delta = (due_date - now).days
    return 0 <= delta <= days


def _normalize_borrow_request_lines(lines: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for line in lines:
        item_name = _to_str(line.get("item_name")).strip()
        item_model = _to_str(line.get("item_model")).strip()
        requested_qty = _to_int(line.get("requested_qty"))
        note = _to_str(line.get("note"))
        if not item_name:
            raise ValueError("item_name is required")
        if not item_model:
            raise ValueError("item_model is required")
        if requested_qty <= 0:
            raise ValueError("requested_qty must be greater than 0")
        normalized.append(
            {
                "item_name": item_name,
                "item_model": item_model,
                "requested_qty": requested_qty,
                "note": note,
            }
        )
    if not normalized:
        raise ValueError("request_lines is required")
    return normalized


def _validate_borrow_request_dates(*, borrow_date_value: Any, due_date_value: Any) -> None:
    borrow_date_raw = _to_str(borrow_date_value).strip()
    due_date_raw = _to_str(due_date_value).strip()
    if not borrow_date_raw:
        raise ValueError("borrow_date is required")
    if not due_date_raw:
        raise ValueError("due_date is required")

    borrow_date = _parse_request_date(borrow_date_raw)
    if borrow_date is None:
        raise ValueError("invalid borrow_date format")
    due_date = _parse_request_date(due_date_raw)
    if due_date is None:
        raise ValueError("invalid due_date format")

    if due_date < borrow_date:
        raise ValueError("due_date cannot be earlier than borrow_date")
    if (due_date - borrow_date).days > MAX_BORROW_RESERVATION_DAYS:
        raise ValueError(f"borrow reservation cannot exceed {MAX_BORROW_RESERVATION_DAYS} days")


def _build_borrow_allocation_stats(
    *,
    allocation_rows: list[dict[str, Any]],
    legacy_item_rows: list[dict[str, Any]] | None = None,
) -> tuple[dict[int, int], dict[int, int]]:
    legacy_rows = legacy_item_rows or []
    request_allocations: dict[int, int] = {}
    line_allocations: dict[int, int] = {}
    for row in allocation_rows:
        request_id = _to_int(row.get("request_id"))
        line_id = _to_int(row.get("line_id"))
        if request_id > 0:
            request_allocations[request_id] = request_allocations.get(request_id, 0) + 1
        if line_id > 0:
            line_allocations[line_id] = line_allocations.get(line_id, 0) + 1
    for row in legacy_rows:
        request_id = _to_int(row.get("request_id"))
        item_id = _to_int(row.get("item_id"))
        if request_id > 0 and item_id > 0:
            request_allocations[request_id] = request_allocations.get(request_id, 0) + 1
    return request_allocations, line_allocations


def _build_borrow_requested_qty_by_request(line_rows: list[dict[str, Any]]) -> dict[int, int]:
    requested_by_request: dict[int, int] = {}
    for row in line_rows:
        request_id = _to_int(row.get("request_id"))
        if request_id <= 0:
            continue
        requested_by_request[request_id] = requested_by_request.get(request_id, 0) + _to_int(row.get("requested_qty"))
    return requested_by_request


def _build_borrow_reservation_usage(
    *,
    request_rows: list[dict[str, Any]],
    line_rows: list[dict[str, Any]],
    allocation_rows: list[dict[str, Any]],
    legacy_item_rows: list[dict[str, Any]] | None = None,
    exclude_request_id: int | None = None,
    today: date | None = None,
) -> dict[tuple[str, str], int]:
    legacy_rows = legacy_item_rows or []
    active_request_ids = {
        _to_int(row.get("id"))
        for row in request_rows
        if _derive_borrow_status(
            due_date_value=row.get("due_date"),
            return_date_value=row.get("return_date"),
            status_value=row.get("status"),
            borrow_date_value=row.get("borrow_date"),
            has_allocations=any(_to_int(alloc.get("request_id")) == _to_int(row.get("id")) for alloc in allocation_rows)
            or any(_to_int(item.get("request_id")) == _to_int(row.get("id")) for item in legacy_rows),
            today=today,
        )
        == "reserved"
    }
    if exclude_request_id is not None:
        active_request_ids.discard(exclude_request_id)

    usage: dict[tuple[str, str], int] = {}
    for line in line_rows:
        request_id = _to_int(line.get("request_id"))
        if request_id not in active_request_ids:
            continue
        key = _borrow_key(line.get("item_name"), line.get("item_model"))
        usage[key] = usage.get(key, 0) + _to_int(line.get("requested_qty"))
    return usage


def _build_inventory_available_by_key(inventory_rows: list[dict[str, Any]]) -> dict[tuple[str, str], int]:
    available: dict[tuple[str, str], int] = {}
    for row in inventory_rows:
        if not _is_blank(row.get("deleted_at")):
            continue
        if _to_str(row.get("asset_status")).strip() != "0":
            continue
        key = _borrow_key(row.get("name"), row.get("model"))
        available[key] = available.get(key, 0) + 1
    return available


def _build_inventory_candidates_by_key(inventory_rows: list[dict[str, Any]]) -> dict[tuple[str, str], list[dict[str, Any]]]:
    candidates: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in inventory_rows:
        if not _is_blank(row.get("deleted_at")):
            continue
        if _to_str(row.get("asset_status")).strip() != "0":
            continue
        key = _borrow_key(row.get("name"), row.get("model"))
        candidates.setdefault(key, []).append(row)
    for rows in candidates.values():
        rows.sort(key=lambda item: _to_int(item.get("id")))
    return candidates


def _collect_shortages(
    requested_totals: dict[tuple[str, str], int],
    available_totals: dict[tuple[str, str], int],
) -> list[dict[str, Any]]:
    shortages: list[dict[str, Any]] = []
    for key, requested_qty in requested_totals.items():
        available_qty = available_totals.get(key, 0)
        if available_qty >= requested_qty:
            continue
        shortages.append(
            {
                "item_name": key[0],
                "item_model": key[1],
                "requested_qty": requested_qty,
                "available_qty": available_qty,
                "shortage_qty": requested_qty - available_qty,
            }
        )
    shortages.sort(key=lambda row: (row["item_name"], row["item_model"]))
    return shortages


def _normalize_borrow_status_in_place(
    row: dict[str, Any],
    *,
    has_allocations: bool = False,
    is_fully_allocated: bool = True,
    today: date | None = None,
) -> bool:
    next_status = _derive_borrow_status(
        due_date_value=row.get("due_date"),
        return_date_value=row.get("return_date"),
        status_value=row.get("status"),
        borrow_date_value=row.get("borrow_date"),
        has_allocations=has_allocations,
        is_fully_allocated=is_fully_allocated,
        today=today,
    )
    previous_status = _to_str(row.get("status")).strip()
    if previous_status == next_status:
        return False
    row["status"] = next_status
    return True


def _to_borrow_api_row(
    row: dict[str, Any],
    *,
    has_allocations: bool = False,
    is_fully_allocated: bool = True,
    today: date | None = None,
) -> dict[str, Any]:
    normalized = dict(row)
    normalized["status"] = _derive_borrow_status(
        due_date_value=row.get("due_date"),
        return_date_value=row.get("return_date"),
        status_value=row.get("status"),
        borrow_date_value=row.get("borrow_date"),
        has_allocations=has_allocations,
        is_fully_allocated=is_fully_allocated,
        today=today,
    )
    normalized["is_due_soon"] = normalized["status"] in {"partial_borrowed", "borrowed", "overdue"} and _is_due_soon(
        due_date_value=row.get("due_date"), return_date_value=row.get("return_date"), today=today, days=3
    )
    return normalized


def _sync_borrow_statuses_in_place(
    *,
    request_rows: list[dict[str, Any]],
    allocation_rows: list[dict[str, Any]],
    line_rows: list[dict[str, Any]] | None = None,
    legacy_item_rows: list[dict[str, Any]] | None = None,
    operation_rows: list[dict[str, Any]] | None = None,
    today: date | None = None,
) -> bool:
    legacy_rows = legacy_item_rows or []
    request_allocations, _ = _build_borrow_allocation_stats(
        allocation_rows=allocation_rows,
        legacy_item_rows=legacy_rows,
    )
    requested_qty_by_request = _build_borrow_requested_qty_by_request(line_rows or [])
    status_changed = False
    for row in request_rows:
        request_id = _to_int(row.get("id"))
        allocated_qty = request_allocations.get(request_id, 0)
        has_allocations = allocated_qty > 0
        requested_qty = requested_qty_by_request.get(request_id, 0)
        is_fully_allocated = not has_allocations or requested_qty <= 0 or allocated_qty >= requested_qty
        previous_status = _to_str(row.get("status")).strip().lower()
        if _normalize_borrow_status_in_place(
            row,
            has_allocations=has_allocations,
            is_fully_allocated=is_fully_allocated,
            today=today,
        ):
            status_changed = True
            next_status = _to_str(row.get("status")).strip().lower()
            if previous_status != "cancelled" and next_status == "cancelled" and operation_rows is not None:
                borrow_date = _parse_request_date(row.get("borrow_date"))
                expired_days = 0
                if borrow_date is not None:
                    current_day = today or date.today()
                    expired_days = max((current_day - borrow_date).days, 0)
                _append_operation_log_entry(
                    operation_rows,
                    action="auto_cancel_reservation",
                    entity="borrow_request",
                    entity_id=request_id if request_id > 0 else None,
                    detail={
                        "request_id": request_id if request_id > 0 else None,
                        "borrower": _to_str(row.get("borrower")),
                        "borrow_date": _to_str(row.get("borrow_date")),
                        "expired_days": expired_days,
                        "rule_days": BORROW_RESERVATION_CANCEL_AFTER_DAYS,
                    },
                )
    return status_changed


def create_borrow_request(request_data: dict[str, Any], request_lines: list[dict[str, Any]]) -> int:
    _validate_borrow_request_dates(
        borrow_date_value=request_data.get("borrow_date"),
        due_date_value=request_data.get("due_date"),
    )
    normalized_lines = _normalize_borrow_request_lines(request_lines)
    with _locked_workbook() as wb:
        request_ws = wb["borrow_requests"]
        operation_ws = wb["operation_logs"]
        legacy_item_rows = _read_rows(wb["borrow_items"])
        line_ws = wb["borrow_request_lines"]
        allocation_ws = wb["borrow_allocations"]
        inventory_ws = wb["inventory_items"]
        operation_rows = _read_rows(operation_ws)
        request_rows = _read_rows(request_ws)
        line_rows = _read_rows(line_ws)
        allocation_rows = _read_rows(allocation_ws)
        inventory_rows = _read_rows(inventory_ws)
        status_changed = _sync_borrow_statuses_in_place(
            request_rows=request_rows,
            allocation_rows=allocation_rows,
            line_rows=line_rows,
            legacy_item_rows=legacy_item_rows,
            operation_rows=operation_rows,
        )

        requested_totals: dict[tuple[str, str], int] = {}
        for line in normalized_lines:
            key = _borrow_key(line["item_name"], line["item_model"])
            requested_totals[key] = requested_totals.get(key, 0) + line["requested_qty"]

        available_totals = _build_inventory_available_by_key(inventory_rows)
        reserved_usage = _build_borrow_reservation_usage(
            request_rows=request_rows,
            line_rows=line_rows,
            allocation_rows=allocation_rows,
            legacy_item_rows=legacy_item_rows,
        )
        reservable_totals = {
            key: max(available_totals.get(key, 0) - reserved_usage.get(key, 0), 0)
            for key in set(available_totals) | set(reserved_usage) | set(requested_totals)
        }
        shortages = _collect_shortages(requested_totals, reservable_totals)
        if shortages:
            raise ValueError(json.dumps({"message": "insufficient_reservable_quantity", "shortages": shortages}, ensure_ascii=False))

        request_id = _next_id(request_rows)
        request_rows.append(
            {
                "id": request_id,
                "borrower": request_data["borrower"],
                "department": request_data["department"],
                "purpose": request_data["purpose"],
                "borrow_date": request_data["borrow_date"],
                "due_date": request_data["due_date"],
                "return_date": "",
                "status": "reserved",
                "memo": request_data["memo"],
                "created_at": _now_str(),
            }
        )
        next_line_id = _next_id(line_rows)
        for index, line in enumerate(normalized_lines):
            line_rows.append(
                {
                    "id": next_line_id + index,
                    "request_id": request_id,
                    "item_name": line["item_name"],
                    "item_model": line["item_model"],
                    "requested_qty": line["requested_qty"],
                    "note": line["note"],
                }
            )
        _write_rows(request_ws, SHEETS["borrow_requests"], request_rows)
        _write_rows(line_ws, SHEETS["borrow_request_lines"], line_rows)
        _write_rows(allocation_ws, SHEETS["borrow_allocations"], allocation_rows)
        _write_rows(inventory_ws, SHEETS["inventory_items"], inventory_rows)
        if status_changed:
            _write_rows(operation_ws, SHEETS["operation_logs"], operation_rows)
        wb.save(DB_PATH)
        return request_id


def list_borrow_requests() -> list[dict[str, Any]]:
    with _locked_workbook() as wb:
        request_ws = wb["borrow_requests"]
        line_rows = _read_rows(wb["borrow_request_lines"])
        operation_ws = wb["operation_logs"]
        legacy_item_rows = _read_rows(wb["borrow_items"])
        allocation_rows = _read_rows(wb["borrow_allocations"])
        operation_rows = _read_rows(operation_ws)
        rows = _read_rows(request_ws)
        status_changed = _sync_borrow_statuses_in_place(
            request_rows=rows,
            allocation_rows=allocation_rows,
            line_rows=line_rows,
            legacy_item_rows=legacy_item_rows,
            operation_rows=operation_rows,
        )
        if status_changed:
            _write_rows(request_ws, SHEETS["borrow_requests"], rows)
            _write_rows(operation_ws, SHEETS["operation_logs"], operation_rows)
            wb.save(DB_PATH)
    request_allocations, _ = _build_borrow_allocation_stats(
        allocation_rows=allocation_rows,
        legacy_item_rows=legacy_item_rows,
    )
    requested_qty_by_request = _build_borrow_requested_qty_by_request(line_rows)
    return sorted(
        (
            _to_borrow_api_row(
                row,
                has_allocations=request_allocations.get(_to_int(row.get("id")), 0) > 0,
                is_fully_allocated=(
                    request_allocations.get(_to_int(row.get("id")), 0) >= requested_qty_by_request.get(_to_int(row.get("id")), 0)
                ),
            )
            for row in rows
        ),
        key=lambda row: _to_int(row.get("id")),
        reverse=True,
    )


def get_borrow_request(request_id: int) -> dict[str, Any] | None:
    with _locked_workbook() as wb:
        request_ws = wb["borrow_requests"]
        line_rows = _read_rows(wb["borrow_request_lines"])
        operation_ws = wb["operation_logs"]
        legacy_item_rows = _read_rows(wb["borrow_items"])
        allocation_rows = _read_rows(wb["borrow_allocations"])
        operation_rows = _read_rows(operation_ws)
        rows = _read_rows(request_ws)
        status_changed = _sync_borrow_statuses_in_place(
            request_rows=rows,
            allocation_rows=allocation_rows,
            line_rows=line_rows,
            legacy_item_rows=legacy_item_rows,
            operation_rows=operation_rows,
        )
        target_row: dict[str, Any] | None = None
        for row in rows:
            if _to_int(row.get("id")) != request_id:
                continue
            target_row = row
            break
        if status_changed:
            _write_rows(request_ws, SHEETS["borrow_requests"], rows)
            _write_rows(operation_ws, SHEETS["operation_logs"], operation_rows)
            wb.save(DB_PATH)
    if target_row is None:
        return None
    request_allocations, _ = _build_borrow_allocation_stats(
        allocation_rows=allocation_rows,
        legacy_item_rows=legacy_item_rows,
    )
    requested_qty_by_request = _build_borrow_requested_qty_by_request(line_rows)
    allocated_qty = request_allocations.get(request_id, 0)
    return _to_borrow_api_row(
        target_row,
        has_allocations=allocated_qty > 0,
        is_fully_allocated=allocated_qty >= requested_qty_by_request.get(request_id, 0),
    )


def list_borrow_items_map(request_ids: set[int] | None = None) -> dict[int, list[dict[str, Any]]]:
    with _locked_workbook() as wb:
        line_rows = _read_rows(wb["borrow_request_lines"])
        allocation_rows = _read_rows(wb["borrow_allocations"])
        legacy_item_rows = _read_rows(wb["borrow_items"])
        inventory_rows = _read_rows(wb["inventory_items"])

    inventory_map = _active_inventory_detail_map(inventory_rows)
    inventory_rows_map = _active_inventory_rows_map(inventory_rows)
    selected_ids = request_ids or set()
    results: dict[int, list[dict[str, Any]]] = {}
    line_map: dict[int, dict[str, Any]] = {}
    for row in line_rows:
        line_id = _to_int(row.get("id"))
        if line_id > 0:
            line_map[line_id] = row

    allocation_by_line_id: dict[int, list[dict[str, Any]]] = {}
    for row in allocation_rows:
        line_id = _to_int(row.get("line_id"))
        allocation_by_line_id.setdefault(line_id, []).append(row)

    for row in line_rows:
        request_id = _to_int(row.get("request_id"))
        if selected_ids and request_id not in selected_ids:
            continue
        line_id = _to_int(row.get("id"))
        allocated_rows = allocation_by_line_id.get(line_id, [])
        allocated_item_ids = [_to_int(alloc.get("item_id")) for alloc in allocated_rows if _to_int(alloc.get("item_id")) > 0]
        results.setdefault(request_id, []).append(
            {
                "id": row.get("id"),
                "request_id": row.get("request_id"),
                "item_id": "",
                "quantity": row.get("requested_qty"),
                "note": row.get("note", ""),
                "item_name": _to_str(row.get("item_name")),
                "item_model": _to_str(row.get("item_model")),
                "requested_qty": _to_int(row.get("requested_qty")),
                "allocated_qty": len(allocated_rows),
                "allocated_item_ids": allocated_item_ids,
            }
        )

    for row in legacy_item_rows:
        request_id = _to_int(row.get("request_id"))
        if selected_ids and request_id not in selected_ids:
            continue
        item_id = _to_int(row.get("item_id"))
        details = inventory_map.get(item_id, {"item_name": None, "item_model": None})
        item_name = _to_str(details.get("item_name")) or _to_str(inventory_rows_map.get(item_id, {}).get("name"))
        item_model = _to_str(details.get("item_model")) or _to_str(inventory_rows_map.get(item_id, {}).get("model"))
        results.setdefault(request_id, []).append(
            {
                "id": row.get("id"),
                "request_id": row.get("request_id"),
                "item_id": row.get("item_id"),
                "quantity": row.get("quantity"),
                "note": row.get("note", ""),
                "item_name": item_name,
                "item_model": item_model,
                "requested_qty": _to_int(row.get("quantity"), default=1),
                "allocated_qty": _to_int(row.get("quantity"), default=1),
                "allocated_item_ids": [item_id] if item_id > 0 else [],
            }
        )

    for grouped_rows in results.values():
        grouped_rows.sort(key=lambda row: _to_int(row.get("id")))
    return results


def list_borrow_items(request_id: int) -> list[dict[str, Any]]:
    return list_borrow_items_map({request_id}).get(request_id, [])


def _validate_pickup_request_context(
    *,
    request_id: int,
    request_rows: list[dict[str, Any]],
    line_rows: list[dict[str, Any]],
    allocation_rows: list[dict[str, Any]],
) -> tuple[dict[str, Any] | None, list[dict[str, Any]]]:
    request_row = next((row for row in request_rows if _to_int(row.get("id")) == request_id), None)
    if request_row is None:
        return None, []
    status = _to_str(request_row.get("status")).strip().lower()
    if status not in {"reserved", "expired", "partial_borrowed"}:
        raise ValueError("only reserved request can be picked up")
    request_lines = sorted(
        (row for row in line_rows if _to_int(row.get("request_id")) == request_id),
        key=lambda row: _to_int(row.get("id")),
    )
    if not request_lines:
        raise ValueError("request_lines is required")
    return request_row, request_lines


def _normalize_scan_code(value: Any) -> str:
    return _to_str(value).strip().lower()


def _inventory_serial_values(row: dict[str, Any]) -> list[str]:
    return [
        _normalize_scan_code(row.get("n_property_sn")),
        _normalize_scan_code(row.get("property_sn")),
        _normalize_scan_code(row.get("n_item_sn")),
        _normalize_scan_code(row.get("item_sn")),
    ]


def list_borrow_pickup_lines(request_id: int) -> list[dict[str, Any]] | None:
    with _locked_workbook() as wb:
        request_ws = wb["borrow_requests"]
        operation_ws = wb["operation_logs"]
        legacy_item_rows = _read_rows(wb["borrow_items"])
        line_rows = _read_rows(wb["borrow_request_lines"])
        allocation_rows = _read_rows(wb["borrow_allocations"])
        inventory_rows = _read_rows(wb["inventory_items"])
        operation_rows = _read_rows(operation_ws)
        request_rows = _read_rows(request_ws)

        status_changed = _sync_borrow_statuses_in_place(
            request_rows=request_rows,
            allocation_rows=allocation_rows,
            line_rows=line_rows,
            legacy_item_rows=legacy_item_rows,
            operation_rows=operation_rows,
        )
        if status_changed:
            _write_rows(request_ws, SHEETS["borrow_requests"], request_rows)
            _write_rows(operation_ws, SHEETS["operation_logs"], operation_rows)
            wb.save(DB_PATH)
        request_row, request_lines = _validate_pickup_request_context(
            request_id=request_id,
            request_rows=request_rows,
            line_rows=line_rows,
            allocation_rows=allocation_rows,
        )
        if request_row is None:
            return None

    candidates_by_key = _build_inventory_candidates_by_key(inventory_rows)
    _, line_allocations = _build_borrow_allocation_stats(
        allocation_rows=allocation_rows,
        legacy_item_rows=legacy_item_rows,
    )
    summaries: list[dict[str, Any]] = []
    for line in request_lines:
        requested_qty = _to_int(line.get("requested_qty"))
        allocated_qty = line_allocations.get(_to_int(line.get("id")), 0)
        remaining_qty = max(requested_qty - allocated_qty, 0)
        if remaining_qty <= 0:
            continue
        key = _borrow_key(line.get("item_name"), line.get("item_model"))
        summaries.append(
            {
                "line_id": _to_int(line.get("id")),
                "item_name": _to_str(line.get("item_name")),
                "item_model": _to_str(line.get("item_model")),
                "requested_qty": requested_qty,
                "allocated_qty": allocated_qty,
                "remaining_qty": remaining_qty,
                "candidate_count": len(candidates_by_key.get(key, [])),
            }
        )
    return summaries


def list_borrow_pickup_line_candidates(
    request_id: int,
    line_id: int,
    *,
    keyword: str = "",
    page: int = 1,
    page_size: int = 50,
) -> dict[str, Any] | None:
    with _locked_workbook() as wb:
        request_ws = wb["borrow_requests"]
        operation_ws = wb["operation_logs"]
        legacy_item_rows = _read_rows(wb["borrow_items"])
        line_rows = _read_rows(wb["borrow_request_lines"])
        allocation_rows = _read_rows(wb["borrow_allocations"])
        inventory_rows = _read_rows(wb["inventory_items"])
        operation_rows = _read_rows(operation_ws)
        request_rows = _read_rows(request_ws)

        status_changed = _sync_borrow_statuses_in_place(
            request_rows=request_rows,
            allocation_rows=allocation_rows,
            line_rows=line_rows,
            legacy_item_rows=legacy_item_rows,
            operation_rows=operation_rows,
        )
        if status_changed:
            _write_rows(request_ws, SHEETS["borrow_requests"], request_rows)
            _write_rows(operation_ws, SHEETS["operation_logs"], operation_rows)
            wb.save(DB_PATH)
        request_row, request_lines = _validate_pickup_request_context(
            request_id=request_id,
            request_rows=request_rows,
            line_rows=line_rows,
            allocation_rows=allocation_rows,
        )
        if request_row is None:
            return None

    target_line = next((row for row in request_lines if _to_int(row.get("id")) == line_id), None)
    if target_line is None:
        raise ValueError("line_id does not belong to request")
    _, line_allocations = _build_borrow_allocation_stats(
        allocation_rows=allocation_rows,
        legacy_item_rows=legacy_item_rows,
    )
    requested_qty = _to_int(target_line.get("requested_qty"))
    allocated_qty = line_allocations.get(line_id, 0)
    remaining_qty = max(requested_qty - allocated_qty, 0)
    if remaining_qty <= 0:
        raise ValueError("line_id is already fully allocated")

    key = _borrow_key(target_line.get("item_name"), target_line.get("item_model"))
    candidates = _build_inventory_candidates_by_key(inventory_rows).get(key, [])
    normalized_keyword = _normalize_scan_code(keyword)
    if normalized_keyword:
        filtered_candidates: list[dict[str, Any]] = []
        for item in candidates:
            serial_values = _inventory_serial_values(item)
            if any(normalized_keyword in serial_value for serial_value in serial_values if serial_value):
                filtered_candidates.append(item)
        candidates = filtered_candidates

    total = len(candidates)
    total_pages = max((total + page_size - 1) // page_size, 1)
    start = (page - 1) * page_size
    end = start + page_size
    paged_candidates = candidates[start:end]
    return {
        "line_id": _to_int(target_line.get("id")),
        "item_name": _to_str(target_line.get("item_name")),
        "item_model": _to_str(target_line.get("item_model")),
        "requested_qty": requested_qty,
        "allocated_qty": allocated_qty,
        "remaining_qty": remaining_qty,
        "items": [
            {
                "id": _to_int(item.get("id")),
                "n_property_sn": _to_str(item.get("n_property_sn")),
                "property_sn": _to_str(item.get("property_sn")),
                "n_item_sn": _to_str(item.get("n_item_sn")),
                "item_sn": _to_str(item.get("item_sn")),
            }
            for item in paged_candidates
        ],
        "page": page,
        "page_size": page_size,
        "total": total,
        "total_pages": total_pages,
    }


def resolve_borrow_pickup_scan(request_id: int, code: str) -> dict[str, Any] | None:
    normalized_code = _normalize_scan_code(code)
    if not normalized_code:
        raise ValueError("scan code is required")

    with _locked_workbook() as wb:
        request_ws = wb["borrow_requests"]
        operation_ws = wb["operation_logs"]
        legacy_item_rows = _read_rows(wb["borrow_items"])
        line_rows = _read_rows(wb["borrow_request_lines"])
        allocation_rows = _read_rows(wb["borrow_allocations"])
        inventory_rows = _read_rows(wb["inventory_items"])
        operation_rows = _read_rows(operation_ws)
        request_rows = _read_rows(request_ws)

        status_changed = _sync_borrow_statuses_in_place(
            request_rows=request_rows,
            allocation_rows=allocation_rows,
            line_rows=line_rows,
            legacy_item_rows=legacy_item_rows,
            operation_rows=operation_rows,
        )
        if status_changed:
            _write_rows(request_ws, SHEETS["borrow_requests"], request_rows)
            _write_rows(operation_ws, SHEETS["operation_logs"], operation_rows)
            wb.save(DB_PATH)
        request_row, request_lines = _validate_pickup_request_context(
            request_id=request_id,
            request_rows=request_rows,
            line_rows=line_rows,
            allocation_rows=allocation_rows,
        )
        if request_row is None:
            return None

    candidates_by_key = _build_inventory_candidates_by_key(inventory_rows)
    _, line_allocations = _build_borrow_allocation_stats(
        allocation_rows=allocation_rows,
        legacy_item_rows=legacy_item_rows,
    )
    eligible_line_ids_by_key: dict[tuple[str, str], list[int]] = {}
    for line in request_lines:
        requested_qty = _to_int(line.get("requested_qty"))
        allocated_qty = line_allocations.get(_to_int(line.get("id")), 0)
        if allocated_qty >= requested_qty:
            continue
        key = _borrow_key(line.get("item_name"), line.get("item_model"))
        eligible_line_ids_by_key.setdefault(key, []).append(_to_int(line.get("id")))

    matched_item: dict[str, Any] | None = None
    matched_key: tuple[str, str] | None = None
    for key, candidates in candidates_by_key.items():
        for item in candidates:
            serial_values = _inventory_serial_values(item)
            if normalized_code in serial_values:
                if matched_item is not None:
                    raise ValueError("scan code matches multiple items")
                matched_item = item
                matched_key = key

    if matched_item is None or matched_key is None:
        raise ValueError("scan code not found")

    eligible_line_ids = eligible_line_ids_by_key.get(matched_key, [])
    if not eligible_line_ids:
        raise ValueError("scanned item does not belong to this request")

    return {
        "item": {
            "id": _to_int(matched_item.get("id")),
            "n_property_sn": _to_str(matched_item.get("n_property_sn")),
            "property_sn": _to_str(matched_item.get("property_sn")),
            "n_item_sn": _to_str(matched_item.get("n_item_sn")),
            "item_sn": _to_str(matched_item.get("item_sn")),
            "item_name": _to_str(matched_item.get("name")),
            "item_model": _to_str(matched_item.get("model")),
        },
        "eligible_line_ids": eligible_line_ids,
    }


def list_borrow_pickup_candidates(request_id: int) -> list[dict[str, Any]] | None:
    with _locked_workbook() as wb:
        request_ws = wb["borrow_requests"]
        operation_ws = wb["operation_logs"]
        legacy_item_rows = _read_rows(wb["borrow_items"])
        line_rows = _read_rows(wb["borrow_request_lines"])
        allocation_rows = _read_rows(wb["borrow_allocations"])
        inventory_rows = _read_rows(wb["inventory_items"])
        operation_rows = _read_rows(operation_ws)
        request_rows = _read_rows(request_ws)

        status_changed = _sync_borrow_statuses_in_place(
            request_rows=request_rows,
            allocation_rows=allocation_rows,
            line_rows=line_rows,
            legacy_item_rows=legacy_item_rows,
            operation_rows=operation_rows,
        )
        if status_changed:
            _write_rows(request_ws, SHEETS["borrow_requests"], request_rows)
            _write_rows(operation_ws, SHEETS["operation_logs"], operation_rows)
            wb.save(DB_PATH)
        request_row, request_lines = _validate_pickup_request_context(
            request_id=request_id,
            request_rows=request_rows,
            line_rows=line_rows,
            allocation_rows=allocation_rows,
        )
        if request_row is None:
            return None

    candidates_by_key = _build_inventory_candidates_by_key(inventory_rows)
    _, line_allocations = _build_borrow_allocation_stats(
        allocation_rows=allocation_rows,
        legacy_item_rows=legacy_item_rows,
    )
    results: list[dict[str, Any]] = []
    for line in request_lines:
        line_id = _to_int(line.get("id"))
        item_name = _to_str(line.get("item_name"))
        item_model = _to_str(line.get("item_model"))
        requested_qty = _to_int(line.get("requested_qty"))
        allocated_qty = line_allocations.get(line_id, 0)
        remaining_qty = max(requested_qty - allocated_qty, 0)
        if remaining_qty <= 0:
            continue
        key = _borrow_key(item_name, item_model)
        candidates = candidates_by_key.get(key, [])
        results.append(
            {
                "line_id": line_id,
                "item_name": item_name,
                "item_model": item_model,
                "requested_qty": requested_qty,
                "allocated_qty": allocated_qty,
                "remaining_qty": remaining_qty,
                "candidates": [
                    {
                        "id": _to_int(item.get("id")),
                        "n_property_sn": _to_str(item.get("n_property_sn")),
                        "property_sn": _to_str(item.get("property_sn")),
                        "n_item_sn": _to_str(item.get("n_item_sn")),
                        "item_sn": _to_str(item.get("item_sn")),
                    }
                    for item in candidates
                ],
            }
        )
    return results


def list_borrow_reservation_options(*, exclude_request_id: int | None = None) -> list[dict[str, Any]]:
    with _locked_workbook() as wb:
        request_ws = wb["borrow_requests"]
        operation_ws = wb["operation_logs"]
        request_rows = _read_rows(request_ws)
        legacy_item_rows = _read_rows(wb["borrow_items"])
        line_rows = _read_rows(wb["borrow_request_lines"])
        allocation_rows = _read_rows(wb["borrow_allocations"])
        inventory_rows = _read_rows(wb["inventory_items"])
        operation_rows = _read_rows(operation_ws)
        status_changed = _sync_borrow_statuses_in_place(
            request_rows=request_rows,
            allocation_rows=allocation_rows,
            line_rows=line_rows,
            legacy_item_rows=legacy_item_rows,
            operation_rows=operation_rows,
        )
        if status_changed:
            _write_rows(request_ws, SHEETS["borrow_requests"], request_rows)
            _write_rows(operation_ws, SHEETS["operation_logs"], operation_rows)
            wb.save(DB_PATH)
    reserved_usage = _build_borrow_reservation_usage(
        request_rows=request_rows,
        line_rows=line_rows,
        allocation_rows=allocation_rows,
        legacy_item_rows=legacy_item_rows,
        exclude_request_id=exclude_request_id,
    )
    available_totals = _build_inventory_available_by_key(inventory_rows)

    combo_keys = {
        _borrow_key(row.get("name"), row.get("model"))
        for row in inventory_rows
        if _is_blank(row.get("deleted_at"))
    }

    options: list[dict[str, Any]] = []
    for combo_key in sorted(combo_keys):
        item_name, item_model = combo_key
        available_qty = available_totals.get(combo_key, 0)
        reserved_qty = reserved_usage.get(combo_key, 0)
        reservable_qty = max(available_qty - reserved_qty, 0)
        options.append(
            {
                "item_name": item_name,
                "item_model": item_model,
                "available_qty": available_qty,
                "reserved_qty": reserved_qty,
                "reservable_qty": reservable_qty,
                "selectable": reservable_qty > 0,
            }
        )
    return options


def update_borrow_request(request_id: int, request_data: dict[str, Any], request_lines: list[dict[str, Any]]) -> bool:
    _validate_borrow_request_dates(
        borrow_date_value=request_data.get("borrow_date"),
        due_date_value=request_data.get("due_date"),
    )
    normalized_lines = _normalize_borrow_request_lines(request_lines)
    with _locked_workbook() as wb:
        request_ws = wb["borrow_requests"]
        operation_ws = wb["operation_logs"]
        legacy_item_rows = _read_rows(wb["borrow_items"])
        line_ws = wb["borrow_request_lines"]
        allocation_ws = wb["borrow_allocations"]
        inventory_ws = wb["inventory_items"]
        operation_rows = _read_rows(operation_ws)
        request_rows = _read_rows(request_ws)
        line_rows = _read_rows(line_ws)
        allocation_rows = _read_rows(allocation_ws)
        inventory_rows = _read_rows(inventory_ws)
        status_changed = _sync_borrow_statuses_in_place(
            request_rows=request_rows,
            allocation_rows=allocation_rows,
            line_rows=line_rows,
            legacy_item_rows=legacy_item_rows,
            operation_rows=operation_rows,
        )

        request_row = next((row for row in request_rows if _to_int(row.get("id")) == request_id), None)
        if request_row is None:
            return False
        if _to_str(request_row.get("status")).strip().lower() not in {"reserved", "expired"}:
            raise ValueError("only reserved request can be updated")
        if any(_to_int(row.get("request_id")) == request_id for row in allocation_rows):
            raise ValueError("allocated borrow request cannot be updated")

        requested_totals: dict[tuple[str, str], int] = {}
        for line in normalized_lines:
            key = _borrow_key(line["item_name"], line["item_model"])
            requested_totals[key] = requested_totals.get(key, 0) + line["requested_qty"]

        available_totals = _build_inventory_available_by_key(inventory_rows)
        reserved_usage = _build_borrow_reservation_usage(
            request_rows=request_rows,
            line_rows=line_rows,
            allocation_rows=allocation_rows,
            legacy_item_rows=legacy_item_rows,
            exclude_request_id=request_id,
        )
        reservable_totals = {
            key: max(available_totals.get(key, 0) - reserved_usage.get(key, 0), 0)
            for key in set(available_totals) | set(reserved_usage) | set(requested_totals)
        }
        shortages = _collect_shortages(requested_totals, reservable_totals)
        if shortages:
            raise ValueError(json.dumps({"message": "insufficient_reservable_quantity", "shortages": shortages}, ensure_ascii=False))

        request_row.update(
            {
                "borrower": request_data["borrower"],
                "department": request_data["department"],
                "purpose": request_data["purpose"],
                "borrow_date": request_data["borrow_date"],
                "due_date": request_data["due_date"],
                "return_date": "",
                "status": "reserved",
                "memo": request_data["memo"],
            }
        )
        line_rows = [row for row in line_rows if _to_int(row.get("request_id")) != request_id]
        next_line_id = _next_id(line_rows)
        for index, line in enumerate(normalized_lines):
            line_rows.append(
                {
                    "id": next_line_id + index,
                    "request_id": request_id,
                    "item_name": line["item_name"],
                    "item_model": line["item_model"],
                    "requested_qty": line["requested_qty"],
                    "note": line["note"],
                }
            )
        _write_rows(request_ws, SHEETS["borrow_requests"], request_rows)
        _write_rows(line_ws, SHEETS["borrow_request_lines"], line_rows)
        _write_rows(allocation_ws, SHEETS["borrow_allocations"], allocation_rows)
        _write_rows(inventory_ws, SHEETS["inventory_items"], inventory_rows)
        if status_changed:
            _write_rows(operation_ws, SHEETS["operation_logs"], operation_rows)
        wb.save(DB_PATH)
        return True


def pickup_borrow_request(request_id: int, selections_data: list[dict[str, Any]]) -> bool:
    with _locked_workbook() as wb:
        request_ws = wb["borrow_requests"]
        operation_ws = wb["operation_logs"]
        legacy_item_rows = _read_rows(wb["borrow_items"])
        line_ws = wb["borrow_request_lines"]
        allocation_ws = wb["borrow_allocations"]
        inventory_ws = wb["inventory_items"]
        movement_ws = wb["movement_ledger"]
        operation_rows = _read_rows(operation_ws)
        request_rows = _read_rows(request_ws)
        line_rows = _read_rows(line_ws)
        allocation_rows = _read_rows(allocation_ws)
        inventory_rows = _read_rows(inventory_ws)
        movement_rows = _read_rows(movement_ws)

        status_changed = _sync_borrow_statuses_in_place(
            request_rows=request_rows,
            allocation_rows=allocation_rows,
            line_rows=line_rows,
            legacy_item_rows=legacy_item_rows,
            operation_rows=operation_rows,
        )
        request_row = next((row for row in request_rows if _to_int(row.get("id")) == request_id), None)
        if request_row is None:
            return False
        status = _to_str(request_row.get("status")).strip().lower()
        if status not in {"reserved", "expired", "partial_borrowed"}:
            raise ValueError("only reserved request can be picked up")

        target_lines = [row for row in line_rows if _to_int(row.get("request_id")) == request_id]
        if not target_lines:
            raise ValueError("request_lines is required")

        line_map = {_to_int(row.get("id")): row for row in target_lines if _to_int(row.get("id")) > 0}
        requested_qty_by_line = {_to_int(row.get("id")): _to_int(row.get("requested_qty")) for row in target_lines}
        _, existing_allocations_by_line = _build_borrow_allocation_stats(
            allocation_rows=allocation_rows,
            legacy_item_rows=legacy_item_rows,
        )
        remaining_qty_by_line = {
            line_id: max(requested_qty_by_line.get(line_id, 0) - existing_allocations_by_line.get(line_id, 0), 0)
            for line_id in requested_qty_by_line
        }
        target_line_ids = set(line_map.keys())
        if not selections_data:
            raise ValueError("pickup selections are required")

        selections_by_line: dict[int, list[int]] = {}
        selected_item_ids: set[int] = set()
        total_selected = 0
        for row in selections_data:
            line_id = _to_int(row.get("line_id"))
            if line_id not in target_line_ids:
                raise ValueError(f"line_id {line_id} does not belong to request")
            if line_id in selections_by_line:
                raise ValueError(f"line_id {line_id} duplicated in pickup selections")
            raw_item_ids = row.get("item_ids")
            if not isinstance(raw_item_ids, list):
                raise ValueError(f"item_ids for line_id {line_id} must be a list")
            line_item_ids: list[int] = []
            line_seen: set[int] = set()
            for raw_item_id in raw_item_ids:
                item_id = _to_int(raw_item_id)
                if item_id <= 0:
                    raise ValueError(f"invalid item_id in line_id {line_id}")
                if item_id in line_seen:
                    raise ValueError(f"item_id {item_id} duplicated in line_id {line_id}")
                if item_id in selected_item_ids:
                    raise ValueError(f"item_id {item_id} duplicated across lines")
                line_seen.add(item_id)
                selected_item_ids.add(item_id)
                line_item_ids.append(item_id)
            selections_by_line[line_id] = line_item_ids
            total_selected += len(line_item_ids)
            remaining_qty = remaining_qty_by_line.get(line_id, 0)
            if remaining_qty <= 0:
                raise ValueError(f"line_id {line_id} is already fully allocated")
            if len(line_item_ids) > remaining_qty:
                raise ValueError(f"line_id {line_id} exceeds remaining qty {remaining_qty}")

        if total_selected <= 0:
            raise ValueError("pickup selections are required")

        inventory_map = _active_inventory_rows_map(inventory_rows)
        next_allocation_id = _next_id(allocation_rows)
        created_allocations: list[dict[str, Any]] = []
        for line in sorted(target_lines, key=lambda row: _to_int(row.get("id"))):
            line_id = _to_int(line.get("id"))
            key = _borrow_key(line.get("item_name"), line.get("item_model"))
            selected_ids = selections_by_line.get(line_id, [])
            if not selected_ids:
                continue
            for item_id in selected_ids:
                inventory_row = _validate_item_status(inventory_map=inventory_map, item_id=item_id, allowed_statuses={"0"})
                item_key = _borrow_key(inventory_row.get("name"), inventory_row.get("model"))
                if item_key != key:
                    raise ValueError(f"item_id {item_id} does not match line_id {line_id}")
                _set_inventory_status_with_movement(
                    movement_rows=movement_rows,
                    row=inventory_row,
                    item_id=item_id,
                    status="2",
                    action="pickup",
                    entity="borrow_request",
                    entity_id=request_id,
                )
                created_allocations.append(
                    {
                        "id": next_allocation_id + len(created_allocations),
                        "request_id": request_id,
                        "line_id": line_id,
                        "item_id": item_id,
                        "note": _to_str(line.get("note")),
                    }
                )

        allocation_rows.extend(created_allocations)
        request_allocations_after_pickup, _ = _build_borrow_allocation_stats(
            allocation_rows=allocation_rows,
            legacy_item_rows=legacy_item_rows,
        )
        requested_qty_by_request = _build_borrow_requested_qty_by_request(target_lines)
        allocated_qty_after_pickup = request_allocations_after_pickup.get(request_id, 0)
        request_row["status"] = _derive_borrow_status(
            due_date_value=request_row.get("due_date"),
            return_date_value="",
            status_value="borrowed",
            borrow_date_value=request_row.get("borrow_date"),
            has_allocations=True,
            is_fully_allocated=allocated_qty_after_pickup >= requested_qty_by_request.get(request_id, 0),
        )
        request_row["return_date"] = ""

        _write_rows(request_ws, SHEETS["borrow_requests"], request_rows)
        _write_rows(line_ws, SHEETS["borrow_request_lines"], line_rows)
        _write_rows(allocation_ws, SHEETS["borrow_allocations"], allocation_rows)
        _write_rows(inventory_ws, SHEETS["inventory_items"], inventory_rows)
        _write_rows(movement_ws, SHEETS["movement_ledger"], movement_rows)
        if status_changed:
            _write_rows(operation_ws, SHEETS["operation_logs"], operation_rows)
        wb.save(DB_PATH)
        return True


def return_borrow_request(request_id: int, *, return_date_value: Any = None) -> bool:
    with _locked_workbook() as wb:
        request_ws = wb["borrow_requests"]
        operation_ws = wb["operation_logs"]
        legacy_item_rows = _read_rows(wb["borrow_items"])
        line_ws = wb["borrow_request_lines"]
        allocation_ws = wb["borrow_allocations"]
        inventory_ws = wb["inventory_items"]
        movement_ws = wb["movement_ledger"]
        operation_rows = _read_rows(operation_ws)
        request_rows = _read_rows(request_ws)
        line_rows = _read_rows(line_ws)
        allocation_rows = _read_rows(allocation_ws)
        inventory_rows = _read_rows(inventory_ws)
        movement_rows = _read_rows(movement_ws)

        status_changed = _sync_borrow_statuses_in_place(
            request_rows=request_rows,
            allocation_rows=allocation_rows,
            line_rows=line_rows,
            legacy_item_rows=legacy_item_rows,
            operation_rows=operation_rows,
        )
        request_row = next((row for row in request_rows if _to_int(row.get("id")) == request_id), None)
        if request_row is None:
            return False
        current_status = _to_str(request_row.get("status")).strip().lower()
        if current_status not in {"partial_borrowed", "borrowed", "overdue"}:
            raise ValueError("only borrowed request can be returned")

        target_allocations = [row for row in allocation_rows if _to_int(row.get("request_id")) == request_id]
        target_item_ids = [_to_int(row.get("item_id")) for row in target_allocations if _to_int(row.get("item_id")) > 0]
        if not target_item_ids:
            legacy_allocations = [
                row for row in legacy_item_rows if _to_int(row.get("request_id")) == request_id and _to_int(row.get("item_id")) > 0
            ]
            target_item_ids = [_to_int(row.get("item_id")) for row in legacy_allocations]
        if not target_item_ids:
            raise ValueError("borrow request has no allocated items")

        inventory_map = _active_inventory_rows_map(inventory_rows)
        for item_id in target_item_ids:
            row = _validate_item_status(inventory_map=inventory_map, item_id=item_id, allowed_statuses={"2"})
            _set_inventory_status_with_movement(
                movement_rows=movement_rows,
                row=row,
                item_id=item_id,
                status="0",
                action="return",
                entity="borrow_request",
                entity_id=request_id,
            )

        normalized_return_date = _parse_request_date(return_date_value) or date.today()
        request_row["return_date"] = normalized_return_date.strftime("%Y/%m/%d")
        request_row["status"] = "returned"

        _write_rows(request_ws, SHEETS["borrow_requests"], request_rows)
        _write_rows(line_ws, SHEETS["borrow_request_lines"], line_rows)
        _write_rows(allocation_ws, SHEETS["borrow_allocations"], allocation_rows)
        _write_rows(inventory_ws, SHEETS["inventory_items"], inventory_rows)
        _write_rows(movement_ws, SHEETS["movement_ledger"], movement_rows)
        if status_changed:
            _write_rows(operation_ws, SHEETS["operation_logs"], operation_rows)
        wb.save(DB_PATH)
        return True


def delete_borrow_request(request_id: int) -> bool:
    with _locked_workbook() as wb:
        request_ws = wb["borrow_requests"]
        item_ws = wb["borrow_items"]
        line_ws = wb["borrow_request_lines"]
        allocation_ws = wb["borrow_allocations"]
        inventory_ws = wb["inventory_items"]
        movement_ws = wb["movement_ledger"]
        request_rows = _read_rows(request_ws)
        item_rows = _read_rows(item_ws)
        line_rows = _read_rows(line_ws)
        allocation_rows = _read_rows(allocation_ws)
        inventory_rows = _read_rows(inventory_ws)
        movement_rows = _read_rows(movement_ws)
        remaining_requests = [row for row in request_rows if _to_int(row.get("id")) != request_id]
        deleted = len(remaining_requests) != len(request_rows)
        if not deleted:
            return False
        request_row = next((row for row in request_rows if _to_int(row.get("id")) == request_id), None)
        old_items = [row for row in item_rows if _to_int(row.get("request_id")) == request_id]
        old_allocations = [row for row in allocation_rows if _to_int(row.get("request_id")) == request_id]
        inventory_map = _active_inventory_rows_map(inventory_rows)
        request_status = ""
        if request_row is not None:
            request_status = _derive_borrow_status(
                due_date_value=request_row.get("due_date"),
                return_date_value=request_row.get("return_date"),
                status_value=request_row.get("status"),
                borrow_date_value=request_row.get("borrow_date"),
                has_allocations=bool(old_allocations or old_items),
            )
        if request_row is not None and _borrow_status_uses_inventory(request_status):
            for allocation in old_allocations:
                item_id = _to_int(allocation.get("item_id"))
                row = _validate_item_status(inventory_map=inventory_map, item_id=item_id, allowed_statuses={"2"})
                _set_inventory_status_with_movement(
                    movement_rows=movement_rows,
                    row=row,
                    item_id=item_id,
                    status="0",
                    action="delete",
                    entity="borrow_request",
                    entity_id=request_id,
                )
            for old_item in old_items:
                item_id = _to_int(old_item.get("item_id"))
                row = _validate_item_status(inventory_map=inventory_map, item_id=item_id, allowed_statuses={"2"})
                _set_inventory_status_with_movement(
                    movement_rows=movement_rows,
                    row=row,
                    item_id=item_id,
                    status="0",
                    action="delete",
                    entity="borrow_request",
                    entity_id=request_id,
                )
        remaining_items = [row for row in item_rows if _to_int(row.get("request_id")) != request_id]
        remaining_lines = [row for row in line_rows if _to_int(row.get("request_id")) != request_id]
        remaining_allocations = [row for row in allocation_rows if _to_int(row.get("request_id")) != request_id]
        _write_rows(request_ws, SHEETS["borrow_requests"], remaining_requests)
        _write_rows(item_ws, SHEETS["borrow_items"], remaining_items)
        _write_rows(line_ws, SHEETS["borrow_request_lines"], remaining_lines)
        _write_rows(allocation_ws, SHEETS["borrow_allocations"], remaining_allocations)
        _write_rows(inventory_ws, SHEETS["inventory_items"], inventory_rows)
        _write_rows(movement_ws, SHEETS["movement_ledger"], movement_rows)
        wb.save(DB_PATH)
        return True


def _active_inventory_map(inventory_rows: list[dict[str, Any]]) -> dict[int, dict[str, Any]]:
    return {
        _to_int(row.get("id")): row
        for row in inventory_rows
        if _is_blank(row.get("deleted_at"))
    }


def _create_issue_request_locked(
    *,
    request_rows: list[dict[str, Any]],
    item_rows: list[dict[str, Any]],
    request_data: dict[str, Any],
    items: list[dict[str, Any]],
) -> int:
    request_id = _next_id(request_rows)
    request_rows.append(
        {
            "id": request_id,
            "requester": request_data["requester"],
            "department": request_data["department"],
            "purpose": request_data["purpose"],
            "request_date": request_data["request_date"],
            "memo": request_data["memo"],
            "created_at": _now_str(),
        }
    )
    next_item_id = _next_id(item_rows)
    for index, item in enumerate(items):
        item_rows.append(
            {
                "id": next_item_id + index,
                "request_id": request_id,
                "item_id": item["item_id"],
                "quantity": item["quantity"],
                "note": item.get("note", ""),
            }
        )
    return request_id


def _create_borrow_request_locked(
    *,
    request_rows: list[dict[str, Any]],
    item_rows: list[dict[str, Any]],
    request_data: dict[str, Any],
    items: list[dict[str, Any]],
) -> int:
    next_status = _derive_borrow_status(
        due_date_value=request_data.get("due_date"),
        return_date_value=request_data.get("return_date"),
    )
    request_id = _next_id(request_rows)
    request_rows.append(
        {
            "id": request_id,
            "borrower": request_data["borrower"],
            "department": request_data["department"],
            "purpose": request_data["purpose"],
            "borrow_date": request_data["borrow_date"],
            "due_date": request_data["due_date"],
            "return_date": request_data["return_date"],
            "status": next_status,
            "memo": request_data["memo"],
            "created_at": _now_str(),
        }
    )
    next_item_id = _next_id(item_rows)
    for index, item in enumerate(items):
        item_rows.append(
            {
                "id": next_item_id + index,
                "request_id": request_id,
                "item_id": item["item_id"],
                "quantity": item["quantity"],
                "note": item.get("note", ""),
            }
        )
    return request_id
